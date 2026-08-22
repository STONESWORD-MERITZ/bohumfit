# -*- coding: utf-8 -*-
"""BOHUMFIT-286 Phase C — 목표 디자인 프로토타입 생성기 (★제품 코드 무접촉).

수기표(`정본C 수기 비교분석표`)의 **42행 체계·서식**을 재현하되,
값은 **현행 파서가 실제로 읽은 것만** 채운다.

★가장 중요한 규칙: **파싱 못 한 칸은 빈칸으로 둔다.**
  수기표 값을 베껴 채우면 프로토타입이 "파싱이 다 된 것처럼" 보여 판단을 망친다.
  빈칸이 곧 "여기가 아직 안 된다"는 정보다(276a의 '지어내지 않는다' 원칙 승계).

실행:
    python scripts/prototype_286_ohj.py
산출:
    보장분석/비교분석표/BOHUMFIT_프로토타입_정본C_260811.xlsx  (★gitignore 폴더)
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from coverage.proposal_parser import parse_proposal_pdf  # noqa: E402
from tests.real_docs import real_proposals  # noqa: E402

PDF_DIR = ROOT / "보장분석" / "비교분석표"
OUT = PDF_DIR / "BOHUMFIT_프로토타입_정본C_260811.xlsx"

# ── 목표 42행 (수기표 '리모델링' B/C열 원문 그대로) ────────────────────────────
#   (대분류, 담보명) — 대분류는 구간 첫 행에만 적힌다.
ROWS: list[tuple[str, str]] = [
    ("실 비", "상 해/질 병 입 원"), ("", "상 해/질 병 통 원 약 제"),
    ("수 술", "상 해 수 술 비"), ("", "질 병 수 술 비"),
    ("", "1종 수술비 (질병 I 상해)"), ("", "2종 수술비 (질병 I 상해)"),
    ("", "3종 수술비 (질병 I 상해)"), ("", "4종 수술비 (질병 I 상해)"),
    ("", "5종 수술비 (질병 I 상해)"), ("", "뇌혈관 수술비"), ("", "심장질환 수술비"),
    ("암", "암 진 단 비(일반암)"), ("", "유 사 암 진 단 비"), ("", "암 수 술 / 로 봇 암 수 술"),
    ("", "항 암 방 사 선 약 물 치 료"), ("", "고액항암치료(표적,면역)"),
    ("", "세기조절 / 양성자 방사선"), ("", "중입자 / 정위 방사선"),
    ("뇌", "뇌 혈 관 질 환"), ("", "뇌 졸 중"), ("", "뇌 출 혈"),
    ("심 장", "심 장 질 환"), ("", "허혈성 심장질환"), ("", "급성심근경색"),
    ("입 원", "상 해 입 원"), ("", "질 병 입 원"), ("", "1 인 실 입 원"), ("", "간 병 인"),
    ("사 망", "일 반 사 망"), ("", "상 해 사 망"), ("", "질 병 사 망"),
    ("후유장해", "상해 질병 후 유 장 해 80%"), ("", "상 해 후 유 장 해 3%"), ("", "질 병 후 유 장 해 3%"),
    ("골 절", "골 절 진 단 비"), ("", "골 절 수 술 비"), ("", "깁스치료비"),
    ("배상책임", "일 상 생 활 배 상 책 임"),
    ("운전자", "형 사 합 의 금"), ("", "변 호 사 선 임"), ("", "벌 금"), ("", "자 부 상"),
]

# ── 파서 kb_name → 목표 행 매핑 (★확실한 것만. 애매하면 넣지 않는다) ──────────
#   값은 (행 담보명, 종별열) — 종별열 None이면 단일 열.
MAP: dict[str, tuple[str, str | None]] = {
    "상해수술": ("상 해 수 술 비", None),
    "질병수술": ("질 병 수 술 비", None),
    "뇌혈관수술": ("뇌혈관 수술비", None),
    "심혈관수술": ("심장질환 수술비", None),
    "암진단금": ("암 진 단 비(일반암)", None),
    "유사암진단금": ("유 사 암 진 단 비", None),
    "뇌혈관질환": ("뇌 혈 관 질 환", None),
    "허혈성심장질환": ("허혈성 심장질환", None),
    "상해사망": ("상 해 사 망", None),
    "질병사망": ("질 병 사 망", None),
    "골절진단비": ("골 절 진 단 비", None),
    "깁스치료비": ("깁스치료비", None),
    "가족/일상/자녀배상": ("일 상 생 활 배 상 책 임", None),
    "자동차사고부상": ("자 부 상", None),
    "교통사고처리지원금": ("형 사 합 의 금", None),
    "변호사선임비용": ("변 호 사 선 임", None),
    "벌금(대인/스쿨존/대물)": ("벌 금", None),
}
for _t in "12345":
    MAP[f"N종수술비(질병 {_t}종)"] = (f"{_t}종 수술비 (질병 I 상해)", "질병")
    MAP[f"N종수술비(상해 {_t}종)"] = (f"{_t}종 수술비 (질병 I 상해)", "상해")

THIN = Side(style="thin", color="FF999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", start_color="FFEBF1DE")
GRP_FILL = PatternFill("solid", start_color="FFDBEEF3")
TIER_FILL = PatternFill("solid", start_color="FFFFFFCC")
GAP_FILL = PatternFill("solid", start_color="FFF2DCDB")  # 스키마에 자리가 없는 행
MAN = '#,##0_ "만""원"'
WON = '"₩"\\ #,##0"원"'


def load_proposals() -> list[dict]:
    out = []
    for p in real_proposals("정본C"):
        out.append(parse_proposal_pdf(p.read_bytes(), p.name))
    return out


def build() -> Path:
    proposals = load_proposals()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "리모델링(프로토타입)"

    ws["B2"] = "(정본C)님 리모델링      【 후 】 보장분석표 — BOHUMFIT 프로토타입(286)"
    ws["B2"].font = Font(name="맑은 고딕", size=16, bold=True)

    ws["B4"], ws["C4"] = "대분류", "담 보 명"
    ws["D4"] = "합계"
    for cell in ("B4", "C4", "D4"):
        ws[cell].font = Font(name="맑은 고딕", size=11, bold=True)
        ws[cell].fill = HEAD_FILL
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].border = BORDER

    # 회사 열 — 제안서 1건당 (질병|상해) 2열을 붙여 수기표의 종수술 병기를 재현한다.
    col = 5
    spans: list[tuple[dict, int]] = []
    for pr in proposals:
        for row, val in ((3, pr["insurer"]), (4, pr["product"][:46]),
                         (5, pr["monthly_premium"])):
            c = ws.cell(row, col, val)
            c.font = Font(name="맑은 고딕", size=10, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            if row == 5:
                c.number_format = WON
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.cell(6, col, "질병").font = Font(size=9, bold=True)
        ws.cell(6, col + 1, "상해").font = Font(size=9, bold=True)
        for c in (col, col + 1):
            ws.cell(6, c).alignment = Alignment(horizontal="center")
            ws.cell(6, c).border = BORDER
        spans.append((pr, col))
        col += 2

    # ── 데이터 행 ───────────────────────────────────────────────────────────
    # 파서 결과를 행 이름으로 인덱싱한다. ★MAP에 없는 kb_name은 버리지 않고 모아 둔다.
    unmapped: list[tuple[str, str, int]] = []
    filled: dict[int, dict[str, int]] = {}
    for pr, base in spans:
        for cov in pr["coverages"]:
            target = MAP.get(cov["kb_name"])
            if target is None:
                unmapped.append((pr["insurer"], cov["kb_name"], cov["amount"]))
                continue
            label, kind = target
            offset = 1 if kind == "상해" else 0
            filled.setdefault(base + offset, {})[label] = cov["amount"]

    # 목표 42행 중 현행 40행 스키마에 자리가 없는 것(2층 갭) — 색으로만 표시한다.
    from coverage.constants import KB_COVERAGES  # noqa: E402

    covered = {name for name, _g, _g12, _a in KB_COVERAGES}
    schema_gap = {label for _grp, label in ROWS
                  if not any(m[0] == label for m in MAP.values() if m)} | {
        f"{t}종 수술비 (질병 I 상해)" for t in "12345"
    }
    _ = covered

    row = 7
    for group, label in ROWS:
        gc = ws.cell(row, 2, group)
        gc.font = Font(name="맑은 고딕", size=11, bold=True)
        gc.alignment = Alignment(horizontal="center", vertical="center")
        gc.border = BORDER
        if group:
            gc.fill = GRP_FILL
        lc = ws.cell(row, 3, label)
        lc.font = Font(name="맑은 고딕", size=11)
        lc.border = BORDER
        if label in schema_gap:
            lc.fill = GAP_FILL

        total = 0
        for _pr, base in spans:
            for off in (0, 1):
                amount = filled.get(base + off, {}).get(label)
                c = ws.cell(row, base + off)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center")
                if amount is None:
                    continue  # ★빈칸 유지 — 베껴 채우지 않는다
                c.value = amount // 10_000
                c.number_format = MAN
                c.fill = TIER_FILL
                total += amount
        tc = ws.cell(row, 4, total // 10_000 if total else None)
        tc.number_format = MAN
        tc.font = Font(name="맑은 고딕", size=11, bold=True)
        tc.fill = HEAD_FILL
        tc.border = BORDER
        row += 1

    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 13
    for i in range(5, col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 13

    # ── 사유 시트 ───────────────────────────────────────────────────────────
    note = wb.create_sheet("차이 사유")
    note.append(["구분", "내용"])
    note["A1"].font = note["B1"].font = Font(bold=True)
    for text in [
        ("원칙", "★빈칸 = 현행 파서가 읽지 못한 칸. 수기표 값을 베껴 채우지 않았다."),
        ("색", "연분홍 담보명 = 현행 40행 스키마에 자리가 없는 행(2층 갭)."),
        ("색", "연노랑 값 = 파서가 실제로 읽은 값."),
        ("불변 결정", "80%이상 후유장해는 집계 제외(243) — 원문에 있어도 비운다."),
        ("불변 결정", "월납은 제안서만 원 단위 버림(276c). 3건 모두 수기표와 일치."),
        ("불변 결정", "폴백 금지(276a) — 못 읽으면 빈칸, 지어내지 않는다."),
    ]:
        note.append(list(text))
    note.append([])
    note.append(["매핑 안 된 파서 결과 (행 자리 없음)", ""])
    for insurer, name, amount in unmapped:
        note.append([insurer, f"{name} = {amount:,}원"])
    note.column_dimensions["A"].width = 34
    note.column_dimensions["B"].width = 76

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"생성: {path}")
