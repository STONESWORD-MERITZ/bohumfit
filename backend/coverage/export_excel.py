"""BOHUMFIT-291(S3)·296: 52행 수기표 양식 4시트 엑셀 — 표지(세로) · 컨설팅 전 · 컨설팅 후 · 최종.

181 다운로드 동선(엔드포인트·build_workbook_bytes 시그니처)은 무파손 유지하고, 산출 구조를
신판 수기표(46행 + 290 확정 3행 + 296 확정 3행 = 52행 · 대분류 11)로 전면 교체한다(248 비분양식 3시트 폐기).

확정 결정(286-D Q1~Q9 · 289 · 290 · 291):
- ★값 계층 동결 — S2(aggregator) 집계 결과(`summary`·`by_company`·`columns`·`stage_totals`)를
  **그대로 옮겨 적는다**. export는 어떤 값도 재계산·보정하지 않는다(값 기입 방식 유지 — 수식 아님).
- 시트명 `컨설팅 전`/`컨설팅 후`(+표지·최종) · 행명은 신판 수기표 문자열 그대로(`중 입 자 치료` 등).
- 회사 열 = 계약당 **2열**(수기표 구조): 종수술 5행·간병인 = 질병|상해. 그 밖은 병합 1값.
  종별을 잃은 값(238 표준환산 → unspecified)은 열로 못 갈라 병합 셀에 합계(추측 금지).
- Q2: 80% 행 값 표시 + 대분류 합계 제외 + "합계 미포함" 셀 메모. Q4: 비고행 블록. Q5: Y/N은
  운전자/배상/실비 행 안(셀 메모·금액 없는 가입 계약은 "Y") — 별도 Y/N 블록 없음.
- `최종`: 기존/점검 후/기대효과 + ★L 접두(케스케이드 하위 행 — 뇌·심장 + 암 계열 · 이 시트만)
  + 종합 판정 블록 = 케스케이드 체인 1:1(compute_stage_totals) + 비고 블록.
- 서식: 구조는 수기표, **색·폰트는 FIT 브랜드 우선**(에메랄드 헤더/그린 티 대분류 선두/라임 특수).
  단위: 보장금액 만원 · 보험료 원. 인쇄 fitToWidth=1.
- overview(합계형) 문서: 대분류 행에 귀속 없는 행이 있으면 합계만(259 가드 · 비고행은 가드 제외).

※ PII: 생성물은 응답 스트림 전용, 서버 미저장.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from openpyxl.comments import Comment

from .aggregator import compute_stage_totals, compute_yn_flags
from .compare import ensure_comparison
from .constants import (
    CASCADE_CHILD_PREFIX_V2,
    KB_COVERAGES_V2,
    LEGACY_TO_V2,
    SHEET_NAME_AFTER_V2,
    SHEET_NAME_BEFORE_V2,
    SHEET_NAME_FINAL_V2,
    SUM_EXCLUDED_NOTE_V2,
    YN_ITEMS_V2,
)
from .v2_mapping import COLUMN_DISEASE, COLUMN_INJURY, COLUMN_UNSPECIFIED, GROUP_APPENDIX_V2, KIND_ROW, ROW_INDEX, resolve
from .excel_style import (
    AMBER_TX,
    BORDER_GRID,
    EMERALD,
    EMERALD_SOFT,
    GRAY_SOFT,
    GRAY_TX,
    GREENTEA,
    INK,
    LIME,
    WHITE,
)

MAN = 10_000
# BOHUMFIT-261 P2: 상담 임팩트용 '20년 동일 기준' 총납입 환산 개월(12개월 × 20년).
#   실제 납만기 기준 총납입(paid_total)과는 성격이 다른 별개 지표다 — 대체가 아니라 병기.
MONTHS_20Y = 240

# BOHUMFIT-291: 라임 특수 강조 행(250 SPECIAL_ITEMS의 V2 대응) — 순환계 치료비. 암 주요치료비는 행이 아니다(분배·S4).
SPECIAL_ROW_IDS: frozenset[str] = frozenset({"circulatory_treatment"})

# BOHUMFIT-250: 양식 원색 → FIT 브랜드 치환(비분양식 강조 "위치"는 원본 실측 그대로).
#   헤더(원본 연노랑) = 에메랄드 면 + 흰 글자 / 강조 행 = 그린 티 면 / 특수 = 라임 면 /
#   시트3 라벨 기본 = 에메랄드 소프트 / [후]·차액 빨강 → 에메랄드(개선)·앰버(악화) 텍스트.
FORM_YELLOW = EMERALD          # 헤더 면(라벨 셀) — 흰 글자와 짝
FORM_BLUE = EMERALD_SOFT       # 시트3 라벨 기본 면
FORM_GRAY = GRAY_SOFT
RED_TX = EMERALD               # 개선 강조 텍스트(FIT 팔레트에 빨강 부재 — 250 S0 근거)
_BORDER = BORDER_GRID


def _company_label(co: dict, companies: list) -> str:
    """BOHUMFIT-240 P1: 계약 라벨 = 회사명, 동일 회사 복수 계약은 '회사명 (n)'."""
    insurer = co.get("insurer")
    if not insurer:
        return f"계약 {co.get('idx')}"
    same = [c for c in companies if c.get("insurer") == insurer]
    if len(same) <= 1:
        return insurer
    ordinal = [str(c.get("idx")) for c in same].index(str(co.get("idx"))) + 1
    return f"{insurer} ({ordinal})"


def _man(value):
    """원 → 만원(정수 반올림). None은 None 유지(미가입 공란)."""
    if value is None:
        return None
    return round(value / MAN)


def _cell(ws, row, col, value=None, *, bold=False, fill=None, fmt=None, align="center",
          color=None, size=10, border=True, wrap=False, name="돋움"):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    # BOHUMFIT-250: 에메랄드 면 위 텍스트는 흰색(대비 10.7:1 — 브랜드 규칙).
    auto_color = WHITE if fill == EMERALD else (color or INK)
    c.font = Font(name=name, bold=bold, size=size, color=auto_color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = _BORDER
    return c


def _unknown_sum(row: dict, ids: set):
    """실계약 키가 아닌 by_company 값 합(만원 변환 전 원 단위) — 없으면 None(공란)."""
    vals = [v for k, v in (row.get("by_company") or {}).items()
            if k not in ids and isinstance(v, (int, float))]
    return sum(vals) if vals else None


def _special_notes(analysis: dict) -> list[str]:
    notes: list[str] = list(analysis.get("warnings") or [])
    comparison = analysis.get("comparison") or {}
    for caution in comparison.get("cautions") or []:
        message = caution.get("message")
        if message:
            notes.append(message)
    seen: set[str] = set()
    unique: list[str] = []
    for note in notes:
        if note not in seen:
            seen.add(note)
            unique.append(note)
    return unique


# ── 시트1: 표지(세로) — S3 실측(병합 9·문구 원문) ────────────────────────────────
def _sheet_cover(ws, analysis: dict, generated_at=None) -> None:
    """BOHUMFIT-261 P1: 표지 리디자인 — 브랜드 밴드 + 고객명 대제목 + 작성일 + 설계사 블록
    + 하단 고지 문구. PDF 표지(export_pdf._cover_page)와 항목·문구를 맞춘다(자료 일관성).
    ★값·집계와 무관한 표시 계층이며 A4 세로 1장 인쇄에 맞춘다.
    """
    ws.title = "표지(세로)"
    cover = analysis.get("report_cover") or {}
    customer = ((analysis.get("before") or {}).get("customer") or {}).get("name")
    display_name = cover.get("customer_name") or customer or "OOO"
    written = cover.get("written_date") or (generated_at.strftime("%Y-%m-%d") if generated_at else "")

    ws.column_dimensions["A"].width = 2.0
    for col in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 6.4
    for merge in ("B2:P4", "B7:P7", "B8:P9", "B11:P11",
                  "C13:F13", "G13:P13", "C14:F14", "G14:P14",
                  "C15:F15", "G15:P15", "C16:F16", "G16:P16",
                  "C17:F17", "G17:P17", "B20:P22"):
        ws.merge_cells(merge)

    # ── 브랜드 밴드(에메랄드 면 + 흰 글자) ────────────────────────────────
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 14
    band = _cell(ws, 2, 2, "BohumFit  보험핏", size=22, bold=True, fill=EMERALD,
                 align="left", border=False, name="맑은 고딕")
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(3, 17):  # 병합 셀 면 색 채우기(테두리 없는 밴드)
        _cell(ws, 2, col, None, fill=EMERALD, border=False)

    # ── 고객명 대제목 ─────────────────────────────────────────────────────
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 40
    _cell(ws, 7, 2, f"{display_name} 님을 위한", size=20, align="left", border=False, name="맑은 고딕")
    _cell(ws, 8, 2, "보험 보장 분석 리포트", size=30, bold=True, align="left", border=False,
          name="맑은 고딕", color=EMERALD)
    if written:
        _cell(ws, 11, 2, f"작성일  {written}", size=11, align="left", border=False,
              color=GRAY_TX, fmt="@")

    # ── 설계사 정보 블록(라벨 + 기입란) ───────────────────────────────────
    #   제공된 값은 채우고 미제공은 빈 기입란(밑줄)으로 남긴다 — 설계사 수기 보완용.
    planner_fields = (
        ("소속(GA)", cover.get("ga_name")),
        ("설계사명", cover.get("planner_name")),
        ("연락처", cover.get("planner_tel")),
        ("E-MAIL", cover.get("planner_email")),
    )
    _cell(ws, 12, 2, "담당 설계사", size=12, bold=True, align="left", border=False, color=EMERALD)
    for offset, (label, value) in enumerate(planner_fields):
        row = 13 + offset
        ws.row_dimensions[row].height = 22
        _cell(ws, row, 3, label, size=11, bold=True, fill=EMERALD_SOFT, align="left")
        _cell(ws, row, 7, value or "", size=11, align="left", fmt="@")

    # ── 하단 고지 문구(PDF 표지와 동일 취지) ──────────────────────────────
    note = _cell(ws, 20, 2,
                 "고객 설명용 요약 리포트입니다. 실제 보장 내용과 보험금 지급 여부는 각 보험사 "
                 "약관과 증권을 따르며, 본 자료는 보험 모집·중개·상품추천·가입권유를 목적으로 "
                 "하지 않습니다.", size=9, align="left", border=False, color=GRAY_TX, wrap=True)
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:P23"



# ═══════════════════════════════════════════════════════════════════════════
# BOHUMFIT-291(S3)·296 — 52행 수기표 양식: 표지 · 컨설팅 전 · 컨설팅 후 · 최종
#   ★값 계층은 S2(aggregator)에서 동결됐다. 이 모듈은 어떤 값도 재계산·보정하지 않는다 —
#     `summary`·`by_company`·`columns`·`stage_totals`(compute_stage_totals)·premium을 **그대로 옮겨 적는다**.
#   ★구조·행명은 신판 수기표(46행 + 290 확정 3행)를 따르고, 서식은 FIT 브랜드가 우선한다
#     (에메랄드 헤더/그린 티 대분류 선두/라임 특수·Pretendard 대신 셀 폰트는 종전 돋움·맑은 고딕 유지).
# ═══════════════════════════════════════════════════════════════════════════

TITLE_BEFORE = "【 전 】"
TITLE_AFTER = "【 후 】"
DATA_ROW0 = 7          # 담보 52행 시작 행(수기표 r7)
HEADER_ROWS = (2, 3, 4, 5, 6)

#: 291 Human 확정 — 케스케이드 하위 행 표시(`L ` 접두). 뇌·심장 + 암 계열. ★`최종` 시트에만.
CASCADE_CHILD_ROWS: frozenset[str] = frozenset({
    "stroke", "cerebral_hemorrhage", "acute_mi",
    "cancer_surgery_davinci", "cancer_surgery_davinci_specific",
    "cancer_drug_targeted", "cancer_drug_immune",
    "radio_imrt", "radio_proton", "radio_carbon",
})
#: 2열 병기 행의 좌|우 열 — ★BOHUMFIT-292(S4·Phase F · Human 확정): 종수술·간병인 모두 **질병|상해**로 통일
#:   (291의 간병인 상해|질병을 수정). 값은 무변경 — 열 순서만.
DUAL_ORDER: dict[str, tuple[str, str]] = {
    **{f"tier_surgery_{t}": (COLUMN_DISEASE, COLUMN_INJURY) for t in range(1, 6)},
    "caregiver": (COLUMN_DISEASE, COLUMN_INJURY),
}
DUAL_LABEL = {COLUMN_DISEASE: "질병", COLUMN_INJURY: "상해"}
#: ★BOHUMFIT-296(Human 확정): 292 F가 넣은 `2열 병기 (좌 | 우)` 헤더 행·`질병 | 상해` 헤더를 **제거**한다.
#:   행명 `1종 수술비 (질병 I 상해)`가 이미 열 순서를 표현하므로 헤더가 중복이다. 각 칸에는 금액만.
#:   → 컨설팅 전/후 시트도 최종 시트처럼 헤더 없이 `DATA_ROW0 + 순서`로 배치한다.
TRACK_LAST_DATA_ROW = DATA_ROW0 + len(KB_COVERAGES_V2) - 1


def track_row_of(row_id: str) -> int:
    """컨설팅 전/후 시트에서 담보 행의 엑셀 행 번호(2열 헤더 제거 — 최종 시트와 동일 좌표)."""
    return DATA_ROW0 + ROW_INDEX[row_id]
#: Q2 표기(상수 단일 소스 `SUM_EXCLUDED_NOTE_V2`).
YN_ROW_IDS = tuple(row.row_id for row in KB_COVERAGES_V2 if row.yn_source)
_YN_ITEM_OF_ROW: dict[str, str] = {}
for _item, _sources in YN_ITEMS_V2:
    for _src in _sources:
        _rid = LEGACY_TO_V2.get(_src)
        if _rid:
            _YN_ITEM_OF_ROW.setdefault(_rid, _item)


def _lead_rows(row_of) -> dict[str, int]:
    """대분류 **선두 강조(GREENTEA)를 받을 행** — ★BOHUMFIT-302b: 회색 헤더(`sum_excluded`)가 대분류
    맨 위에 오면 그 행은 회색이므로, 선두 강조는 **회색이 아닌 첫 행**이 받는다(대분류 구분 유지).
    회색 행뿐인 대분류는 없으므로 폴백은 필요 없다."""
    lead: dict[str, int] = {}
    for spec in KB_COVERAGES_V2:
        if spec.sum_excluded:
            continue
        lead.setdefault(spec.group, row_of(spec.row_id))
    return lead


def _row_by_id(before_like: dict) -> dict:
    """행 조회(row_id 기준). ★구 페이로드(row_id 없음 — 과거 저장분·구 클라이언트)는 이름을 `resolve()`로
    V2 행에 **투영**해 읽는다(값 계산 없음 · 표시 전용 · 같은 행에 둘 이상 오면 먼저 온 것)."""
    rows = {}
    for row in (before_like or {}).get("coverages", []):
        if row.get("row_id"):
            rows[row["row_id"]] = row
    for row in (before_like or {}).get("coverages", []):
        if row.get("row_id"):
            continue
        target = resolve(row.get("kb_name"))
        if target.kind == KIND_ROW and target.row_id not in rows:
            rows[target.row_id] = row
    return rows


def _appendix_rows(before_like: dict) -> list[dict]:
    out = []
    for row in (before_like or {}).get("coverages", []):
        if not row.get("enrolled"):
            continue
        if row.get("group12") == GROUP_APPENDIX_V2:
            out.append(row)
        elif not row.get("row_id") and resolve(row.get("kb_name")).kind != KIND_ROW:
            out.append(row)  # 구 페이로드의 미매칭 행도 비고 블록에 보존
    return out


def _company_columns_available(before_like: dict) -> bool:
    """BOHUMFIT-259 → 291: 회사 열 전개 가드 — 대분류 행(비고 제외) 중 귀속 없는 overview 행이 있으면 합계만."""
    rows = [
        row for row in (before_like or {}).get("coverages", [])
        if row.get("overview") and row.get("enrolled") and row.get("group12") != GROUP_APPENDIX_V2
    ]
    if not rows:
        return True
    return all(any(v is not None for v in (row.get("by_company") or {}).values()) for row in rows)


def _unknown_bucket_present(before_like: dict, companies: list) -> bool:
    ids = {str(c.get("idx")) for c in companies}
    for row in (before_like or {}).get("coverages", []):
        if not row.get("enrolled") or row.get("group12") == GROUP_APPENDIX_V2:
            continue
        if any(k not in ids and v is not None for k, v in (row.get("by_company") or {}).items()):
            return True
    return False


def _yn_by_row(before_like: dict) -> dict:
    """행 → (Y/N, {계약: 'Y'}) — Q5: Y/N은 운전자/배상/실비 행 안에 표시(별도 블록 없음)."""
    coverages = (before_like or {}).get("coverages", [])
    flags = {f["item"]: f for f in compute_yn_flags(coverages)} if coverages else {}
    out = {}
    for row_id in YN_ROW_IDS:
        item = _YN_ITEM_OF_ROW.get(row_id)
        flag = flags.get(item) or {}
        out[row_id] = (flag.get("value", "N"), flag.get("by_company") or {})
    return out


def _dual_cells(row: dict, order: tuple[str, str], key: str | None):
    """2열 병기 셀 값 (좌, 우, 병합값). key=None이면 합계(summary), 아니면 계약 키.
    ★종별을 잃은 값(`unspecified`)이 있으면 열로 못 갈라 **병합 셀 하나**에 행 값을 싣는다(추측 금지)."""
    columns = row.get("columns") or {}

    def _v(col):
        cell = columns.get(col) or {}
        return cell.get("summary") if key is None else (cell.get("by_company") or {}).get(key)

    unspecified = _v(COLUMN_UNSPECIFIED)
    if unspecified is not None:
        merged = row.get("summary") if key is None else (row.get("by_company") or {}).get(key)
        return None, None, merged
    return _v(order[0]), _v(order[1]), None


def _sheet_track(ws, analysis: dict, before_like: dict, *, is_after: bool) -> None:
    """`컨설팅 전` / `컨설팅 후` — 52행 × (합계 2열 + 회사별 2열). 수기표 r2~r6 머리행·r7~ 담보행 구조."""
    ws.title = SHEET_NAME_AFTER_V2 if is_after else SHEET_NAME_BEFORE_V2
    customer = (before_like.get("customer") or {}).get("name") or ""
    companies = (before_like.get("contract_list") or before_like.get("companies") or []) \
        if _company_columns_available(before_like) else []
    n = len(companies)
    ids = {str(c.get("idx")) for c in companies}
    unk = 1 if (n and _unknown_bucket_present(before_like, companies)) else 0
    has_proposal = any((co.get("consulting_status") == "신규제안") or co.get("is_proposal") for co in companies)
    new_slot = 1 if (is_after and n and not has_proposal) else 0

    # 열: A 여백 | B 대분류 | C~E 담보명 | F~G 합계 | 회사별 2열씩 | (미확인 2열) | (신규 골격 2열)
    col_grp, col_name0, col_sum = 2, 3, 6
    col_co0 = 8
    slots = n + unk + new_slot
    col_end = col_co0 + 2 * slots  # exclusive
    ws.column_dimensions["A"].width = 1.6
    ws.column_dimensions["B"].width = 9.0
    for c, w in ((3, 5.5), (4, 13.0), (5, 5.5), (6, 8.5), (7, 8.5)):
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(col_co0, col_end):
        ws.column_dimensions[get_column_letter(c)].width = 8.5

    def _pair(row, col, value=None, **kw):
        """좌우 2열을 병합해 한 값 — 비병기 행·머리행용."""
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        _cell(ws, row, col, value, **kw)
        _cell(ws, row, col + 1, None, **{k: v for k, v in kw.items() if k in ("fill", "fmt", "border")})

    # ── 머리행 r2~r6 ────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=col_grp, end_row=6, end_column=col_name0 + 2)
    _cell(ws, 2, col_grp, f"{customer or 'OOO'}님 리모델링 {TITLE_AFTER if is_after else TITLE_BEFORE}\n보장분석표",
          bold=True, size=13, fill=EMERALD, wrap=True, fmt="@")
    ws.merge_cells(start_row=2, start_column=col_sum, end_row=3, end_column=col_sum + 1)
    _cell(ws, 2, col_sum, "회사명 / 납기만기\n보험기간", bold=True, fill=EMERALD_SOFT, wrap=True, fmt="@")
    ws.merge_cells(start_row=4, start_column=col_sum, end_row=5, end_column=col_sum + 1)
    _cell(ws, 4, col_sum, "상 품 명", bold=True, fill=EMERALD_SOFT, fmt="@")
    _pair(6, col_sum, (before_like.get("premium") or {}).get("monthly_total"), bold=True,
          fill=EMERALD_SOFT, fmt='"₩" #,##0"원"')
    for r, h in ((2, 30), (3, 24), (4, 26), (5, 20), (6, 22)):
        ws.row_dimensions[r].height = h

    def _period(co: dict) -> str:
        return f"{co.get('pay_years') or '-'}년납/{co.get('maturity') or '-'}"

    for i, co in enumerate(companies):
        col = col_co0 + 2 * i
        _pair(2, col, _company_label(co, companies), bold=True, fill=EMERALD, wrap=True, fmt="@")
        # 236·254 정보 보존: 구분(유지/해지·납입완료)·계피관계·가입일은 수기표에 자리가 없어 **셀 메모**로 남긴다.
        status = co.get("consulting_status") or "유지"
        if co.get("paid_up"):
            status += "(납입완료)"
        remark = co.get("remark") or ""
        if "계피상이" in remark:
            status += "·계피상이"
        elif remark and "계피" in remark:
            status += "·계피동일"
        ws.cell(row=2, column=col).comment = Comment(
            "구분: " + status + chr(10) + "가입일: " + str(co.get("contract_date") or "-"), "BohumFit")
        _pair(3, col, _period(co), wrap=True, size=9, fmt="@")
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col + 1)
        _cell(ws, 4, col, co.get("product") or "-", wrap=True, size=8, fmt="@")
        _cell(ws, 4, col + 1, None); _cell(ws, 5, col, None); _cell(ws, 5, col + 1, None)
        _pair(6, col, co.get("monthly_premium"), fmt='"₩" #,##0"원"')
    if unk:
        col = col_co0 + 2 * n
        _pair(2, col, "계약 미확인", bold=True, fill=EMERALD, wrap=True, fmt="@")
        _pair(3, col, "(원문 계약 특정 불가)", size=8, wrap=True, fmt="@")
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col + 1)
        _cell(ws, 4, col, "-", fmt="@"); _cell(ws, 4, col + 1, None); _cell(ws, 5, col, None); _cell(ws, 5, col + 1, None)
        _pair(6, col, None, fmt="#,##0")
    if new_slot:
        col = col_co0 + 2 * (n + unk)
        _pair(2, col, "신규 설계 반영 대상", bold=True, fill=EMERALD, wrap=True, fmt="@")
        _pair(3, col, "(가입제안서 반영 전)", size=8, wrap=True, fmt="@")
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col + 1)
        _cell(ws, 4, col, "-", fmt="@"); _cell(ws, 4, col + 1, None); _cell(ws, 5, col, None); _cell(ws, 5, col + 1, None)
        _pair(6, col, None, fmt="#,##0")

    # ── 담보 52행 ───────────────────────────────────────────────────────
    rows_by_id = _row_by_id(before_like)
    yn = _yn_by_row(before_like)
    group_start: dict[str, int] = {}
    group_end: dict[str, int] = {}
    lead_row = _lead_rows(track_row_of)
    for spec in KB_COVERAGES_V2:
        row = track_row_of(spec.row_id)  # ★296: 2열 헤더 제거 — 헤더 행 삽입 없음
        data = rows_by_id.get(spec.row_id) or {}
        group_start.setdefault(spec.group, row)
        group_end[spec.group] = row
        lead = lead_row.get(spec.group) == row
        # ★BOHUMFIT-302b: `sum_excluded` 행은 **회색 헤더**로 구분한다(합계에 안 더해지는 행임을 면으로 알린다).
        #   FIT 팔레트의 `GRAY_SOFT`를 쓰고, 대분류 선두(GREENTEA)·특수 강조(LIME)보다 우선한다.
        row_fill = GRAY_SOFT if spec.sum_excluded else (
            GREENTEA if lead else (LIME if spec.row_id in SPECIAL_ROW_IDS else None))
        ws.row_dimensions[row].height = 18
        # 담보명(C~E 병합) — ★수기표 문자열 그대로. Q2·Q5 표기는 셀 메모.
        ws.merge_cells(start_row=row, start_column=col_name0, end_row=row, end_column=col_name0 + 2)
        label = _cell(ws, row, col_name0, spec.display, bold=True, fill=EMERALD_SOFT, wrap=True, fmt="@")
        _cell(ws, row, col_name0 + 1, None, fill=EMERALD_SOFT); _cell(ws, row, col_name0 + 2, None, fill=EMERALD_SOFT)
        notes = []
        if spec.sum_excluded:
            notes.append(f"{SUM_EXCLUDED_NOTE_V2}(243) — 값은 표시하되 대분류 합계에 더하지 않음")
        if spec.row_id in yn:
            value, per_co = yn[spec.row_id]
            notes.append(f"가입특약 Y/N: {value}")
        if notes:
            label.comment = Comment("\n".join(notes), "BohumFit")
        if spec.dual_column:
            order = DUAL_ORDER[spec.row_id]
            left, right, merged = _dual_cells(data, order, None)
            if merged is not None:
                _pair(row, col_sum, _man(merged), bold=True, fmt="#,##0", fill=row_fill)
            else:
                _cell(ws, row, col_sum, _man(left), bold=True, fmt="#,##0", fill=row_fill)
                _cell(ws, row, col_sum + 1, _man(right), bold=True, fmt="#,##0", fill=row_fill)
            for i, co in enumerate(companies):
                col = col_co0 + 2 * i
                left, right, merged = _dual_cells(data, order, str(co.get("idx")))
                if merged is not None:
                    _pair(row, col, _man(merged), fmt="#,##0", fill=row_fill)
                else:
                    _cell(ws, row, col, _man(left), fmt="#,##0", fill=row_fill)
                    _cell(ws, row, col + 1, _man(right), fmt="#,##0", fill=row_fill)
        else:
            _pair(row, col_sum, _man(data.get("summary")), bold=True, fmt="#,##0", fill=row_fill)
            per_co = yn.get(spec.row_id, ("N", {}))[1] if spec.row_id in yn else {}
            for i, co in enumerate(companies):
                key = str(co.get("idx"))
                amount = (data.get("by_company") or {}).get(key)
                value = _man(amount)
                if value is None and per_co.get(key) == "Y":
                    value = "Y"  # Q5: 금액 없이 가입만 확인되는 계약(합계-only)
                _pair(row, col_co0 + 2 * i, value, fmt="#,##0" if not isinstance(value, str) else "@", fill=row_fill)
        if unk:
            _pair(row, col_co0 + 2 * n, _man(_unknown_sum(data, ids)), fmt="#,##0", fill=row_fill)
        if new_slot:
            _pair(row, col_co0 + 2 * (n + unk), None, fmt="#,##0", fill=row_fill)
    last_data_row = TRACK_LAST_DATA_ROW
    # 대분류(B) 병합 — 그룹 구간
    for group, start in group_start.items():
        end = group_end[group]
        if end > start:
            ws.merge_cells(start_row=start, start_column=col_grp, end_row=end, end_column=col_grp)
        _cell(ws, start, col_grp, group, bold=True, fill=GREENTEA, wrap=True, fmt="@")
        for r in range(start + 1, end + 1):
            _cell(ws, r, col_grp, None, fill=GREENTEA)

    # ★296: 2열 병기 헤더 행 제거(행명이 이미 `(질병 I 상해)` 순서 표현). 첫 종수술 행 라벨 메모로만 안내.
    ws.cell(row=track_row_of("tier_surgery_1"), column=col_name0).comment = Comment(
        "2열 병기: 좌=질병 · 우=상해 (종수술 1~5종 · 간병인 동일). 종별 미확인(표준환산)은 병합 셀에 합계.", "BohumFit")

    # ── 비고 블록 ───────────────────────────────────────────────────────
    appendix = _appendix_rows(before_like)
    row = last_data_row + 1
    ws.merge_cells(start_row=row, start_column=col_grp, end_row=row + max(len(appendix), 1) - 1, end_column=col_grp)
    _cell(ws, row, col_grp, "비고", bold=True, fill=GRAY_SOFT, fmt="@")
    if not appendix:
        ws.merge_cells(start_row=row, start_column=col_name0, end_row=row, end_column=col_end - 1)
        _cell(ws, row, col_name0, "-", fmt="@")
        last_row = row
    else:
        for offset, data in enumerate(appendix):
            r = row + offset
            if offset:
                _cell(ws, r, col_grp, None, fill=GRAY_SOFT)
            ws.merge_cells(start_row=r, start_column=col_name0, end_row=r, end_column=col_name0 + 2)
            _cell(ws, r, col_name0, data.get("kb_name"), fill=GRAY_SOFT, wrap=True, fmt="@", align="left")
            _pair(r, col_sum, _man(data.get("summary")), bold=True, fmt="#,##0")
            for i, co in enumerate(companies):
                _pair(r, col_co0 + 2 * i, _man((data.get("by_company") or {}).get(str(co.get("idx")))), fmt="#,##0")
            if unk:
                _pair(r, col_co0 + 2 * n, _man(_unknown_sum(data, ids)), fmt="#,##0")
            if new_slot:
                _pair(r, col_co0 + 2 * (n + unk), None, fmt="#,##0")
        last_row = row + len(appendix) - 1

    # 하단 안내
    note_row = last_row + 2
    ws.merge_cells(start_row=note_row, start_column=col_grp, end_row=note_row, end_column=max(col_end - 1, col_sum + 1))
    note = "보장금액 단위: 만원 / 보험료: 원 · 종수술·간병인은 2열 병기(질병|상해) · 후유장해 80%는 대분류 합계 미포함(243)"
    if any((rows_by_id.get(r) or {}).get("estimated") for r in rows_by_id):
        note += " · 표준환산 종수술은 종별 미확인으로 병합 표기"
    _cell(ws, note_row, col_grp, note, align="left", border=False, color=GRAY_TX, fmt="@", size=9)

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(max(col_end - 1, col_sum + 1))}{note_row}"
    ws.freeze_panes = ws.cell(row=DATA_ROW0, column=col_co0)


def _final_label(spec) -> str:
    return f"{CASCADE_CHILD_PREFIX_V2}{spec.display}" if spec.row_id in CASCADE_CHILD_ROWS else spec.display


def _sheet_final(ws, analysis: dict, before: dict, after_before: dict | None) -> None:
    """`최종` — 기존/점검 후/기대효과 3열 + L 접두 + 종합 판정 17행 블록 + 비고 블록."""
    ws.title = SHEET_NAME_FINAL_V2
    customer = (before.get("customer") or {}).get("name") or ""
    b_rows, a_rows = _row_by_id(before), _row_by_id(after_before) if after_before else {}
    b_prem = (before.get("premium") or {}).get("monthly_total")
    a_prem = (after_before.get("premium") or {}).get("monthly_total") if after_before else None
    prem_delta = (a_prem - b_prem) if (a_prem is not None and b_prem is not None) else None
    prem_delta_color = EMERALD if (prem_delta is None or prem_delta <= 0) else AMBER_TX

    col_grp, col_name0, col_b, col_a, col_d = 2, 3, 6, 8, 10
    ws.column_dimensions["A"].width = 1.6
    ws.column_dimensions["B"].width = 9.0
    for c, w in ((3, 5.5), (4, 13.0), (5, 5.5), (6, 8.5), (7, 8.5), (8, 8.5), (9, 8.5), (10, 8.5), (11, 8.5)):
        ws.column_dimensions[get_column_letter(c)].width = w

    def _pair(row, col, value=None, **kw):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        _cell(ws, row, col, value, **kw)
        _cell(ws, row, col + 1, None, **{k: v for k, v in kw.items() if k in ("fill", "fmt", "border")})

    ws.merge_cells(start_row=2, start_column=col_grp, end_row=6, end_column=col_name0 + 2)
    _cell(ws, 2, col_grp, f"{customer or 'OOO'}님 리모델링 【 최종 】\n보장분석표", bold=True, size=13,
          fill=EMERALD, wrap=True, fmt="@")
    for col, title in ((col_b, "기존"), (col_a, "점검 후"), (col_d, "기대효과")):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col + 1)
        _cell(ws, 2, col, title, bold=True, fill=EMERALD, fmt="@")
        _cell(ws, 2, col + 1, None, fill=EMERALD); _cell(ws, 3, col, None, fill=EMERALD); _cell(ws, 3, col + 1, None, fill=EMERALD)
    # r4 월납 · r5 여백 라벨 · r6 20년 총납입(월납×240 — 261 P2 표기 지표)
    _pair(4, col_b, b_prem, bold=True, fill=EMERALD_SOFT, fmt='"₩" #,##0"원"')
    _pair(4, col_a, a_prem, bold=True, fill=EMERALD_SOFT, fmt='"₩" #,##0"원"')
    _pair(4, col_d, prem_delta, bold=True, fill=EMERALD_SOFT, fmt='"₩" #,##0"원"', color=prem_delta_color)
    _pair(5, col_b, "월납 보험료", size=8, fmt="@"); _pair(5, col_a, "월납 보험료", size=8, fmt="@"); _pair(5, col_d, "차액(후−전)", size=8, fmt="@")
    _pair(6, col_b, b_prem * MONTHS_20Y if b_prem is not None else None, fmt='"₩" #,##0"원"')
    _pair(6, col_a, a_prem * MONTHS_20Y if a_prem is not None else None, fmt='"₩" #,##0"원"')
    _pair(6, col_d, prem_delta * MONTHS_20Y if prem_delta is not None else None,
          fmt='"₩" #,##0"원"', color=prem_delta_color)
    # 261 P2 지표 유지: 20년 납부 시 총납입 차액 = 월납 차액 × 240(동일 기준 환산 — 라벨은 메모로).
    ws.cell(row=6, column=col_d).comment = Comment("20년 납부 시 총납입 차액 = 월납 차액(후−전) × 240", "BohumFit")
    ws.cell(row=6, column=col_b).comment = Comment("20년 납부 시 총납입(월납 × 240)", "BohumFit")
    for r, h in ((2, 30), (3, 18), (4, 22), (5, 14), (6, 20)):
        ws.row_dimensions[r].height = h

    group_start: dict[str, int] = {}
    group_end: dict[str, int] = {}
    lead_row = _lead_rows(lambda rid: DATA_ROW0 + ROW_INDEX[rid])
    for offset, spec in enumerate(KB_COVERAGES_V2):
        row = DATA_ROW0 + offset
        group_start.setdefault(spec.group, row); group_end[spec.group] = row
        lead = lead_row.get(spec.group) == row
        # ★BOHUMFIT-302b: `sum_excluded` 행은 **회색 헤더**로 구분한다(합계에 안 더해지는 행임을 면으로 알린다).
        #   FIT 팔레트의 `GRAY_SOFT`를 쓰고, 대분류 선두(GREENTEA)·특수 강조(LIME)보다 우선한다.
        row_fill = GRAY_SOFT if spec.sum_excluded else (
            GREENTEA if lead else (LIME if spec.row_id in SPECIAL_ROW_IDS else None))
        ws.merge_cells(start_row=row, start_column=col_name0, end_row=row, end_column=col_name0 + 2)
        label = _cell(ws, row, col_name0, _final_label(spec), bold=True, fill=EMERALD_SOFT, wrap=True, fmt="@",
                      align="left" if spec.row_id in CASCADE_CHILD_ROWS else "center")
        _cell(ws, row, col_name0 + 1, None, fill=EMERALD_SOFT); _cell(ws, row, col_name0 + 2, None, fill=EMERALD_SOFT)
        if spec.sum_excluded:
            label.comment = Comment(f"{SUM_EXCLUDED_NOTE_V2}(243)", "BohumFit")
        b_val = (b_rows.get(spec.row_id) or {}).get("summary")
        a_val = (a_rows.get(spec.row_id) or {}).get("summary") if after_before else None
        _pair(row, col_b, _man(b_val), bold=True, fmt="#,##0", fill=row_fill)
        _pair(row, col_a, _man(a_val) if after_before else None, bold=True, fmt="#,##0", fill=row_fill)
        delta = (a_val or 0) - (b_val or 0) if after_before and (a_val is not None or b_val is not None) else None
        _pair(row, col_d, _man(delta) if delta is not None else None, fmt="#,##0",
              fill=row_fill, color=(EMERALD if (delta or 0) > 0 else AMBER_TX if (delta or 0) < 0 else None))
    for group, start in group_start.items():
        end = group_end[group]
        if end > start:
            ws.merge_cells(start_row=start, start_column=col_grp, end_row=end, end_column=col_grp)
        _cell(ws, start, col_grp, group, bold=True, fill=GREENTEA, wrap=True, fmt="@")
        for r in range(start + 1, end + 1):
            _cell(ws, r, col_grp, None, fill=GREENTEA)
    row = DATA_ROW0 + len(KB_COVERAGES_V2)

    # ── 종합 판정 블록 — 케스케이드 체인 1:1 (S2 compute_stage_totals) ─────────
    stages_b = compute_stage_totals(before.get("coverages", []))
    stages_a = compute_stage_totals(after_before.get("coverages", [])) if after_before else {}
    row += 1
    ws.merge_cells(start_row=row, start_column=col_grp, end_row=row, end_column=col_d + 1)
    _cell(ws, row, col_grp, "종합 판정 — 진단 시 지급 합계(케스케이드) · 전 / 후 / 증감", bold=True, fill=EMERALD, align="left", fmt="@")
    stage_row0 = row + 1
    for offset, (key, value_b) in enumerate(stages_b.items()):
        r = stage_row0 + offset
        ws.merge_cells(start_row=r, start_column=col_grp, end_row=r, end_column=col_name0 + 2)
        _cell(ws, r, col_grp, key, bold=True, fill=EMERALD_SOFT, align="left", fmt="@")
        _pair(r, col_b, _man(value_b), fmt="#,##0")
        value_a = stages_a.get(key) if after_before else None
        _pair(r, col_a, _man(value_a) if after_before else None, fmt="#,##0")
        delta = (value_a or 0) - (value_b or 0) if after_before else None
        _pair(r, col_d, _man(delta) if delta is not None else None, fmt="#,##0",
              color=(EMERALD if (delta or 0) > 0 else AMBER_TX if (delta or 0) < 0 else None))
    row = stage_row0 + len(stages_b) - 1

    # ── 비고 블록 ───────────────────────────────────────────────────────
    appendix = {}
    for data in _appendix_rows(before):
        appendix.setdefault(data["kb_name"], [None, None])[0] = data.get("summary")
    for data in (_appendix_rows(after_before) if after_before else []):
        appendix.setdefault(data["kb_name"], [None, None])[1] = data.get("summary")
    row += 1
    ws.merge_cells(start_row=row, start_column=col_grp, end_row=row + max(len(appendix), 1) - 1, end_column=col_grp)
    _cell(ws, row, col_grp, "비고", bold=True, fill=GRAY_SOFT, fmt="@")
    if not appendix:
        ws.merge_cells(start_row=row, start_column=col_name0, end_row=row, end_column=col_d + 1)
        _cell(ws, row, col_name0, "-", fmt="@")
    for offset, (name, (bv, av)) in enumerate(appendix.items()):
        r = row + offset
        if offset:
            _cell(ws, r, col_grp, None, fill=GRAY_SOFT)
        ws.merge_cells(start_row=r, start_column=col_name0, end_row=r, end_column=col_name0 + 2)
        _cell(ws, r, col_name0, name, fill=GRAY_SOFT, align="left", wrap=True, fmt="@")
        _pair(r, col_b, _man(bv), fmt="#,##0")
        _pair(r, col_a, _man(av) if after_before else None, fmt="#,##0")
        delta = (av or 0) - (bv or 0) if after_before and (av is not None or bv is not None) else None
        _pair(r, col_d, _man(delta) if delta is not None else None, fmt="#,##0")
    last_row = row + max(len(appendix), 1) - 1

    # 특이사항
    notes = _special_notes(analysis)
    if notes:
        last_row += 2
        ws.merge_cells(start_row=last_row, start_column=col_grp, end_row=last_row, end_column=col_d + 1)
        _cell(ws, last_row, col_grp, "특이사항", bold=True, fill=EMERALD_SOFT, align="left", fmt="@")
        for note in notes:
            last_row += 1
            ws.merge_cells(start_row=last_row, start_column=col_grp, end_row=last_row, end_column=col_d + 1)
            _cell(ws, last_row, col_grp, f"· {note}", align="left", wrap=True, fmt="@", size=9)
    last_row += 2
    ws.merge_cells(start_row=last_row, start_column=col_grp, end_row=last_row, end_column=col_d + 1)
    _cell(ws, last_row, col_grp,
          f"보장금액 단위: 만원 / 보험료: 원 · '{CASCADE_CHILD_PREFIX_V2.strip()}' 접두 = 상위 진단 지급에 더해지는 하위 담보 · "
          f"후유장해 80%는 대분류 합계 미포함(243) · 20년 총납입 = 월납×{MONTHS_20Y}(동일 기준 환산)",
          align="left", border=False, color=GRAY_TX, fmt="@", size=9)

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(col_d + 1)}{last_row}"
    ws.freeze_panes = ws.cell(row=DATA_ROW0, column=col_b)


def build_workbook_bytes(analysis: dict, generated_at=None) -> bytes:
    """분석 dict([전]만 또는 전후 비교 결과) → 52행 수기표 양식 4시트 xlsx 바이트.

    BOHUMFIT-291(S3): 표지(세로) · 컨설팅 전 · 컨설팅 후 · 최종. ★값은 S2 집계 그대로.
    """
    before = analysis.get("before", {}) or {}
    after = analysis.get("after") or {}
    after_before = after.get("before") or None
    try:
        ensure_comparison(analysis)  # comparison 부재 시 파생(특이사항 cautions 노출용)
    except Exception:
        pass  # 비교 파생 실패는 특이사항 축소일 뿐 — 생성 자체는 진행(정보 보존)
    wb = Workbook()
    _sheet_cover(wb.active, analysis, generated_at or datetime.now())
    _sheet_track(wb.create_sheet(), analysis, before, is_after=False)
    # [후]가 없으면 컨설팅 후 시트는 [전]과 동일 구조의 빈 값(양식 유지 — 시트 4개 고정).
    _sheet_track(wb.create_sheet(), analysis, after_before if after_before else {"customer": before.get("customer"), "coverages": [], "premium": {}}, is_after=True)
    _sheet_final(wb.create_sheet(), analysis, before, after_before)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
