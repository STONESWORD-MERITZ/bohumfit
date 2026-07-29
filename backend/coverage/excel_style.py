"""BOHUMFIT-250: 엑셀 산출물 FIT 브랜드 스타일 계층 — 상수·프리셋 단일 소스.

구조·강조 위치는 비분양식 원본 실측(250 S0), 색은 FIT 브랜드(src/index.css 정본)를 따른다.
★면 전용 색(LIME·GREENTEA)은 fill 전용 — 폰트/보더 색으로 사용 금지(테스트로 고정).
값·구조는 건드리지 않는다(export_excel의 값 기입과 독립된 스타일 적용 전용).
"""
from __future__ import annotations

from openpyxl.styles import Border, Font, PatternFill, Side

# FIT 팔레트(openpyxl ARGB — 앞 FF 불투명). 출처: src/index.css(브랜드 정본)·구 exporter FIT v1.1.
EMERALD = "FF084734"        # 에메랄드 파인 — 헤더 면·강조 텍스트(흰 글자 대비 10.7:1)
EMERALD_SOFT = "FFEEF6F1"   # 에메랄드 소프트 면
GREENTEA = "FFCDEDB3"       # 그린 티 — ★면 전용(대분류 선두 강조 행)
LIME = "FFCEF17B"           # 라임 글로우 — ★면 전용(특수 강조: 암 주요치료비·순환계 치료비)
INK = "FF0A0A0A"            # 잉크(제목)
BODY = "FF1E293B"           # 본문
GRAY_SOFT = "FFF1F1F1"
GRAY_TX = "FF7A7A78"
AMBER_TX = "FFB45309"       # 악화/증가 경고 텍스트(구 exporter 상태 색 선례)
WHITE = "FFFFFFFF"

# 면 전용 색 목록 — 코드 검사 테스트가 이 목록으로 "폰트 색 사용 0"을 고정한다.
FILL_ONLY_COLORS = (LIME, GREENTEA)

FILL_HEADER = PatternFill("solid", fgColor=EMERALD)
FILL_HIGHLIGHT = PatternFill("solid", fgColor=GREENTEA)
FILL_SPECIAL = PatternFill("solid", fgColor=LIME)
FILL_SOFT = PatternFill("solid", fgColor=EMERALD_SOFT)
FILL_GRAY = PatternFill("solid", fgColor=GRAY_SOFT)

_thin = Side(style="thin", color="FFB9C4BD")
_medium = Side(style="medium", color=EMERALD)
BORDER_GRID = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_BOX = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)

# BOHUMFIT-254 개정1: 구획 테두리 — 얇은 그리드는 유지하고 블록 경계(회사 열군 ↔ 합계 ↔
#   담보명 ↔ [후])와 섹션 경계에만 굵은 선을 덧댄다. 레이아웃 정본의 테두리 밀도는 styles.xml
#   cellXfs 결손으로 판독 불가(254 S0 기록) — 가독 기준 설계이며 값·구조에는 영향이 없다.
SIDE_THIN = _thin
SIDE_SECTION = _medium


def section_border(*, left: bool = False, right: bool = False,
                   top: bool = False, bottom: bool = False) -> Border:
    """지정한 변만 굵은 구획선, 나머지는 기존 얇은 그리드."""
    return Border(
        left=SIDE_SECTION if left else SIDE_THIN,
        right=SIDE_SECTION if right else SIDE_THIN,
        top=SIDE_SECTION if top else SIDE_THIN,
        bottom=SIDE_SECTION if bottom else SIDE_THIN,
    )


def font(*, bold: bool = False, size: int = 10, color: str = INK, name: str = "맑은 고딕") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


FONT_HEADER = font(bold=True, color=WHITE)
FONT_TITLE = font(bold=True, size=13, color=EMERALD)

# 비분양식 원본 실측(250 S0) — 강조 담보 행(값 셀 포함 행 전체 연노랑 → 그린 티 치환).
HIGHLIGHT_ITEMS = frozenset({
    "일반사망",
    "상해후유장해", "질병후유장해",
    "암진단금", "유사암진단금",
    "뇌혈관질환",
    "심혈관질환",
    "일반종수술 1종(표준환산)", "일반종수술 2종(표준환산)", "일반종수술 3종(표준환산)",
    "일반종수술 4종(표준환산)", "일반종수술 5종(표준환산)",
})
# 원본 특수 강조(파랑 theme4 → 라임 면 치환).
SPECIAL_ITEMS = frozenset({"암 주요치료비", "순환계 치료비"})

# 시트2 우측 고객정보 패널 라벨(원본 O열 실측 — 값은 설계사 수기용 공란 유지·PII 미기입).
PANEL_LABELS: tuple[tuple[int, str, bool], ...] = (
    (2, "고등전산", False),
    (3, "보험나이", False),
    (6, "1.성명 : ", False),
    (7, "2.주민 : ", False),
    (9, "3.휴대 : ", False),
    (10, "4.주소 : ", False),
    (11, "5.직업 : ", False),
    (12, "6.신체 : ", False),
    (13, "7.운전 : ", False),
    (14, "8.병원", False),
    (17, "[양식]", True),
    (22, "손해(건강)", False),
    (23, "20년/20년", False),
    (24, "20년/100세", False),
)
