# -*- coding: utf-8 -*-
"""BOHUMFIT-254 회귀 — 엑셀 개정(테두리 구획·보험료 상단·Y/N 회사별·디자인).

★불변 계약: 담보 금액·회사합=합계·40행·[후] 이월은 그대로. 이 파일은 개정 4항목이
값을 바꾸지 않으면서 적용됐는지만 고정한다. 익명 합성 픽스처(PII 0).
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from coverage.aggregator import build_before, compute_yn_flags
from coverage.compare import build_after_analysis
from coverage.excel_style import EMERALD, SIDE_SECTION
from coverage.export_excel import FORM_ITEMS, YN_ROWS, build_workbook_bytes

MAN = 10_000


# BOHUMFIT-290(S2): 집계 행이 V2 49행 — 구 이름 조회는 export와 같은 투영(legacy_form_view)으로.
from coverage.aggregator import legacy_form_view as _view  # noqa: E402
from coverage.v2_mapping import GROUP_APPENDIX_V2 as _APPENDIX  # noqa: E402


def _raw(with_yn: bool = True) -> dict:
    """계약 2개 — 계약1은 운전자 특약군, 계약2는 실손·자동차부상 보유(회사별 Y 표적)."""
    extra = {}
    if with_yn:
        extra = {
            "벌금(대인/스쿨존/대물)": {"agg": "sum", "by_company": {"1": 300 * MAN}},
            "교통사고처리지원금": {"agg": "sum", "by_company": {"1": 500 * MAN}},
            "자동차사고부상": {"agg": "sum", "by_company": {"2": 100 * MAN}},
            "상해입원의료비": {"agg": "rep", "by_company": {"2": 5000 * MAN}},
        }
    return {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [
            {"idx": 1, "insurer": "합성손보1", "product": "합성상품1호", "contract_date": "2021-01-01",
             "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
             "monthly_premium": 10_000, "remark": "계피상이(계약자 김*순·피보험자 홍길동)"},
            {"idx": 2, "insurer": "합성손보2", "product": "합성상품2호", "contract_date": "2022-01-01",
             "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
             "monthly_premium": 20_000},
        ],
        "matrix": {
            "질병사망": {"by_company": {"1": 3000 * MAN, "2": 2000 * MAN}},
            "암진단금": {"by_company": {"1": 1000 * MAN}},
        },
        "diagnosis": {}, "notes": {}, "extra": extra, "warnings": [],
    }


def _analysis(with_yn: bool = True) -> dict:
    before = build_before(_raw(with_yn), today="2026-07-27")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def _ws(with_yn: bool = True):
    result = build_after_analysis(_analysis(with_yn), {"existing": [], "proposals": []})
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))
    return wb["비교분석표"], result


# ── 개정3: Y/N 회사별 (집계 파생 — 금액 불변) ────────────────────────────────────
def test_yn_flags_expose_per_company_without_changing_amounts():
    """compute_yn_flags가 by_company를 파생하되 원천 금액·합계 규칙은 불변."""
    before = build_before(_raw(), today="2026-07-27")
    rows = _view(before["coverages"])
    flags = {f["item"]: f for f in compute_yn_flags(before["coverages"])}

    assert flags["운전자특약"]["value"] == "Y"
    assert flags["운전자특약"]["by_company"] == {"1": "Y"}          # 계약1만 보유
    assert flags["자동차부상치료비"]["by_company"] == {"2": "Y"}
    assert flags["상해실손의료비"]["by_company"] == {"2": "Y"}
    assert flags["가족일상배상책임"]["value"] == "N"
    assert flags["가족일상배상책임"]["by_company"] == {}            # 미보유 → 공란(키 없음)
    # ★금액 불변: 원천 by_company는 그대로 숫자(플래그가 덮어쓰지 않음).
    assert rows["벌금(대인/스쿨존/대물)"]["by_company"] == {"1": 300 * MAN}
    assert rows["자동차사고부상"]["by_company"] == {"2": 100 * MAN}


def test_sheet2_yn_rows_render_per_company_y():
    """시트2 Y/N 5행 — 보유 계약 열에 "Y", 미보유 열은 공란, 합계 열 규칙 불변."""
    ws, _result = _ws()
    n = 2
    col_bsum = 2 + n
    col_asum = col_bsum + 4
    col_a0 = col_asum + 1
    yn_row = {item: 45 + i for i, item in enumerate(YN_ROWS)}

    row = yn_row["운전자특약"]
    assert ws.cell(row=row, column=2).value == "Y"        # 계약1 보유
    assert ws.cell(row=row, column=3).value is None       # 계약2 미보유 → 공란
    assert ws.cell(row=row, column=col_bsum).value == "Y"  # 합계 규칙 불변
    assert ws.cell(row=row, column=col_a0).value == "Y"    # [후] 이월도 회사별 Y

    row = yn_row["자동차부상치료비"]
    assert ws.cell(row=row, column=2).value is None
    assert ws.cell(row=row, column=3).value == "Y"

    row = yn_row["가족일상배상책임"]
    assert ws.cell(row=row, column=col_bsum).value == "N"
    assert all(ws.cell(row=row, column=c).value is None for c in (2, 3))


def test_yn_recomputed_from_coverages_not_stale_payload():
    """★실사용 동선 결함 고정: 클라이언트 [후] payload는 yn_flags가 [전] 값 그대로
    스프레드된다(coverages만 해지 반영). 시트는 담보 행에서 재파생해야 해지가 반영된다."""
    analysis = _analysis()
    result = build_after_analysis(analysis, {"existing": [], "proposals": []})
    stale = [dict(f) for f in result["before"]["yn_flags"]]      # [전] 스냅샷(운전자=계약1 Y)
    after_before = result["after"]["before"]
    # 계약1 해지 재현: 담보 행에서 계약1 귀속을 제거하고 payload 플래그는 낡은 채로 둔다.
    for row in after_before["coverages"]:
        by_company = {k: v for k, v in (row.get("by_company") or {}).items() if k != "1"}
        row["by_company"] = by_company
        row["enrolled"] = any(v is not None for v in by_company.values())
        row["summary"] = sum(v for v in by_company.values() if v is not None) or None
    after_before["yn_flags"] = stale                            # ★낡은 [전] 플래그
    after_before["contract_list"] = [c for c in after_before["contract_list"] if str(c["idx"]) != "1"]

    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))
    ws = wb["비교분석표"]
    row = 45 + YN_ROWS.index("운전자특약")
    n = 2
    col_asum = (2 + n) + 4
    assert ws.cell(row=row, column=2 + n).value == "Y"          # [전] 합계는 그대로 Y
    assert ws.cell(row=row, column=col_asum).value == "N"       # ★[후]는 해지 반영 N
    assert ws.cell(row=row, column=col_asum + 1).value is None  # [후] 남은 계약2 열도 공란


# ── 개정2: 보험료 상단 이동(값 불변·위치만) ──────────────────────────────────────
def test_premium_moved_to_meta_block_top():
    """9행 "월보험료" = 회사별 월납 + 합계 월납. 하단 50행은 비고 값 중복 0."""
    ws, result = _ws()
    n = 2
    col_bsum = 2 + n
    col_asum = col_bsum + 4
    col_a0 = col_asum + 1
    assert ws.cell(row=9, column=2).value == 10_000
    assert ws.cell(row=9, column=3).value == 20_000
    assert ws.cell(row=9, column=col_bsum).value == (result["before"]["premium"]["monthly_total"])
    assert ws.cell(row=9, column=col_asum).value == (result["after"]["before"]["premium"]["monthly_total"])
    assert ws.cell(row=9, column=col_a0).value == 10_000
    # 라벨은 원본 정본 표기 그대로("보험료 합계") — 위치만 이동, 하단 중복 0.
    assert ws.cell(row=9, column=col_bsum + 1).value == "보험료 합계"
    for col in range(1, ws.max_column + 1):
        assert ws.cell(row=50, column=col).value is None


def test_meta_block_keeps_kp_difference_inline():
    """계피관계 행은 사라지지 않고 "구 분" 행에 병기(정보 보존 — 236 패턴)."""
    ws, _result = _ws()
    assert ws.cell(row=6, column=2).value.endswith("계피상이")
    assert ws.cell(row=6, column=3).value == "유지"
    labels = [ws.cell(row=r, column=2 + 2 + 1).value for r in range(6, 10)]
    assert labels == ["구 분", "가입일", "납만기", "보험료 합계"]


# ── 개정1: 구획 테두리 ─────────────────────────────────────────────────────────
def test_section_borders_applied_on_block_edges():
    """합계 열·담보명 블록 경계는 굵은 선, 담보 내부는 얇은 그리드 유지."""
    ws, _result = _ws()
    n = 2
    col_bsum = 2 + n
    col_name0 = col_bsum + 1
    first_item_row = 10
    assert ws.cell(row=first_item_row, column=col_bsum).border.left.style == SIDE_SECTION.style
    assert ws.cell(row=first_item_row, column=col_name0).border.left.style == SIDE_SECTION.style
    assert ws.cell(row=9, column=col_bsum).border.bottom.style == SIDE_SECTION.style  # 메타 블록 하단
    # 대분류 전환 행 상단 = 섹션 구분선 / 같은 대분류 내부 행은 얇은 선.
    #   사망(일반~상해사망) → 후유장해(상해후유장해~) 전환 지점.
    section_row = 10 + FORM_ITEMS.index("상해후유장해")
    assert ws.cell(row=section_row, column=2).border.top.style == SIDE_SECTION.style
    inner_row = 10 + FORM_ITEMS.index("질병후유장해")   # 같은 대분류 — 줄 없음
    assert ws.cell(row=inner_row, column=2).border.top.style == "thin"
    inner_row2 = 10 + FORM_ITEMS.index("유사암진단금")  # 암 대분류 내부
    assert ws.cell(row=inner_row2, column=2).border.top.style == "thin"
    # 구획선 색은 브랜드 에메랄드(면 전용 색 아님 — 250 계약 유지).
    assert ws.cell(row=first_item_row, column=col_bsum).border.left.color.rgb == EMERALD


def test_amount_cells_unchanged_by_refinements():
    """★값 불변: 담보 금액 셀·회사합=합계가 개정 후에도 그대로."""
    ws, result = _ws()
    rows = _view(result["before"]["coverages"])
    n = 2
    col_bsum = 2 + n
    for item in ("질병사망", "암진단금"):
        row = 10 + FORM_ITEMS.index(item)
        expected = (rows[item].get("by_company") or {})
        cells = []
        for idx in range(n):
            got = ws.cell(row=row, column=2 + idx).value
            exp = expected.get(str(idx + 1))
            assert got == (None if exp is None else round(exp / MAN))
            cells.append(got or 0)
        assert sum(cells) == ws.cell(row=row, column=col_bsum).value
