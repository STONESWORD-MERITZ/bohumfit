from __future__ import annotations

import io
from datetime import datetime

from openpyxl import load_workbook

from coverage.aggregator import build_before
from coverage.export_excel import build_workbook_bytes
from coverage.export_pdf import build_coverage_html


def _analysis() -> dict:
    raw = {
        "customer": {"name": None, "age": None, "sex": None},
        "contracts": [
            {
                "idx": 1,
                "insurer": "메리츠화재",
                "product": "(무) 알파보험",
                "contract_date": "2020-01-01",
                "pay_cycle": "월납",
                "pay_years": 20,
                "pay_months": 240,
                "maturity": "90세",
                "monthly_premium": 573_227,
            },
            {
                "idx": 2,
                "insurer": "DB손보",
                "product": "BOP보험",
                "contract_date": "2021-02-02",
                "pay_cycle": "월납",
                "pay_years": None,
                "pay_months": None,
                "maturity": "100세",
                "monthly_premium": None,
                "remark": "보험료 미제공",
            },
        ],
        "notes": {},
        "matrix": {"상해사망": {"by_company": {"1": 550_000_000, "2": None}}},
        "extra": {},
    }
    before = build_before(raw)
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def test_before_payload_promotes_contract_list_with_missing_premium() -> None:
    before = _analysis()["before"]
    contract = before["contract_list"][0]

    assert contract["insurer"] == "DB손보"
    assert contract["monthly_premium"] is None
    assert contract["maturity"] == "100세"
    assert contract["remark"] == "보험료 미제공"


def test_excel_before_sheet_has_contract_list_block() -> None:
    wb = load_workbook(io.BytesIO(build_workbook_bytes(_analysis())))
    ws = wb["전 회사별세부"]

    assert [ws.cell(3, col).value for col in range(1, 8)] == [
        "번호",
        "회사명",
        "상품명",
        "납입기간",
        "만기",
        "월보험료",
        "비고",
    ]
    assert ws.cell(4, 2).value == "DB손보"
    assert ws.cell(4, 6).value == "미제공"
    assert ws.cell(4, 7).value == "보험료 미제공"


def test_pdf_html_has_contract_list_block() -> None:
    html = build_coverage_html(_analysis(), datetime(2026, 6, 1))

    assert "<th>회사명</th>" in html
    assert '<th class="num">월보험료</th>' in html
    assert "DB손보" in html
    assert "BOP보험" in html
    assert "미제공" in html
    assert "보험료 미제공" in html
