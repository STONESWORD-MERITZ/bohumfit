from __future__ import annotations

import io
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from coverage.compare import build_after_analysis
from coverage.export_excel import build_workbook_bytes
from coverage.export_pdf import build_coverage_html

sys.path.insert(0, os.path.dirname(__file__))
from test_coverage_compare_188 import _analysis, _plan  # noqa: E402


def _report() -> dict:
    return build_after_analysis(_analysis(), _plan())


def test_screen_per_rider_table_focuses_on_coverage_amounts() -> None:
    source = Path(__file__).resolve().parents[2] / "src" / "pages" / "CoverageRemodel.tsx"
    text = source.read_text(encoding="utf-8")
    head = text.split("comparisonGroups.map", 1)[1].split("<tbody>", 1)[0]

    assert "table-fixed" in head
    assert "<colgroup>" in head
    assert "전 보장금액" in head
    assert "후 보장금액" in head
    assert re.search(r">\s*증감\s*</th>", head)
    assert ">상태</th>" not in head
    assert ">변화</th>" not in head


def test_screen_proposal_amount_editor_is_collapsed_by_default() -> None:
    source = Path(__file__).resolve().parents[2] / "src" / "pages" / "CoverageRemodel.tsx"
    text = source.read_text(encoding="utf-8")
    section = text.split('h3 className="ko-heading text-base font-bold text-ink-900">핵심 보장금액', 1)[1].split(
        "{afterResult &&", 1
    )[0]

    assert "expandedProposalIds" in section
    assert "toggleProposalExpanded" in section
    assert 'aria-expanded={expanded}' in section
    assert '{expanded ? "접기" : "펼치기"}' in section
    assert "{expanded && (" in section
    assert "핵심 보장금액을 입력해 주세요." in section


def test_pdf_per_rider_compare_table_focuses_on_coverage_amounts() -> None:
    html = build_coverage_html(_report())
    section = html.split("특약별 보장금액 비교", 1)[1].split("</table>", 1)[0]

    assert "<th>대분류</th>" in section
    assert "<th>담보</th>" in section
    assert '<th class="num">전 보장금액</th>' in section
    assert '<th class="num">후 보장금액</th>' in section
    assert '<th class="num">증감(후−전)</th>' in section

    rows = re.findall(r"<tr><td.*?</tr>", section, re.S)
    row = next(item for item in rows if ">수술비</td>" in item)
    assert row.count("<td") == 5
    assert "+1,000만" in row
    assert "부족" not in row
    assert "충분" not in row


def test_excel_compare_sheet_uses_amount_columns_only() -> None:
    # BOHUMFIT-248 P2: 엑셀 산출을 비분양식 3시트로 정본화 — 검증 의도를 신 양식 등가로 갱신.
    workbook = load_workbook(io.BytesIO(build_workbook_bytes(_report())))
    sheet = workbook["컨설팅 전"]
    values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value is not None]
    # 신 양식: 담보 행은 금액(만원 숫자)만 — 진단 상태 문자열은 시트에 없다.
    assert "부족" not in values and "충분" not in values and "미가입" not in values
    assert "질 병 수 술 비" in values              # 291: V2 행명(수기표 문자열) 존재
