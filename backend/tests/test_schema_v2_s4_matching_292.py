# -*- coding: utf-8 -*-
"""BOHUMFIT-292(S4) — 매칭·분배 배선 계약.

★고정하는 계약
  B  EXTRA_PATTERNS 협소화: 세기조절·양성자·카티·표적·중입자는 `항암약물방사선`에 **안 잡힌다**(각자 라벨/미포섭).
     세기조절·양성자·카티는 매트릭스 표적항암치료 셀에서 배타 차감(target_included · 246 중입자 기전 확장)
  C  다빈치 종별: 특정암제외/일반암 → 일반 · 전립선 · 갑상선 · 특정암 · **미기재 → 종별 미상(needs_review, 일반암 추측 0)**
     KB분류 암수술 라인은 매트릭스 암수술 셀에서 차감(matrix_included) → 행 분리(sum/rep 아님)
  D  통합치료비 자동 판정: 내역 금액 상이 → Q8형 / 동액·미기재 → 주요치료비형. 한도 문구는 항목이 아니다.
     Q8(순환계): 본체 → 순환계 치료비 + 수술 → 뇌·심장 수술비 / 수술 항목 없음 → 빈칸 + needs_review(폴백 금지)
     공존: 직접 특약 + 분배분 → **sum**(origin=direct+distributed) · 직접 특약끼리 → rep · 분배분끼리 → sum
  E  합성 라벨: 약물만 → 항암 약물 치료 · 방사선만 → 방사선 치료 · 결합 담보 → `항암약물방사선` 비고 보존(갈라 추측 0)
     `심·뇌혈관수술비`: 뇌만/심만 → 각자 행 · 둘 다/미상 → 비고 보존
  F  엑셀 2열 헤더(질병 | 상해) 종수술 위 삽입 · 간병인 열 순서 질병|상해 통일 · 최종 시트 좌표 무변경 · 값 무변경
"""
from __future__ import annotations

from pathlib import Path

import pytest

from coverage.aggregator import build_before, build_v2_rows, _apply_exclusions
from coverage.compare import build_after_analysis
from coverage.constants import (
    DAVINCI_UNKNOWN_LABEL,
    KB_COVERAGES_V2,
    TARGET_INCLUDED_LABELS,
    anticancer_label,
    classify_extra,
    davinci_label,
)
from coverage.export_excel import DATA_ROW0, DUAL_ORDER, build_workbook_bytes, track_row_of
from coverage.export_pdf import DUAL_ORDER as PDF_DUAL_ORDER
from coverage.integrated_treatment import (
    MODE_MAJOR,
    MODE_Q8,
    ORIGIN_DISTRIBUTED,
    distribution_entries,
    extract_integrated_treatments,
    route_item,
)
from coverage.proposal_parser import _entry, _merge_entries, parse_proposal_pdf, parse_proposal_text
from coverage.v2_mapping import ROW_INDEX, resolve

MAN = 10_000
REAL = Path(__file__).resolve().parents[2] / "보장분석" / "비교분석표"


def _row(rows, rid):
    return next(c for c in rows if c.get("row_id") == rid)


# ── B. 과포섭 정정 ────────────────────────────────────────────────────────
@pytest.mark.parametrize("line, label", [
    ("37 정액 표적항암방사선치료비(항암세기조절방사선)(최초1회한) 고액항암치료비 1,000만", "세기조절방사선"),
    ("42 정액 항암세기조절방사선치료비(1회한)(간편할인형)(갱신형) 고액항암치료비 500만", "세기조절방사선"),
    ("39 정액 항암양성자방사선치료비(최초1회한) 고액항암치료비 1,000만", "양성자방사선"),
    ("17 정액 갱신형 항암중입자방사선치료비(통합간편가입) 기타 인보험(정액)담보 5,000만", "중입자방사선"),
    ("40 정액 카티(CAR-T)항암약물허가치료비(1회한)(간편할인형)(갱신형) 고액항암치료비 1,000만", "면역항암치료"),
    ("38 정액 표적항암약물허가치료비(최초1회한) 고액항암치료비 3,000만", None),
    ("35 정액 항암방사선치료비 항암방사선약물치료비 1,000만", "항암방사선치료비"),
    ("36 정액 항암약물치료비 항암방사선약물치료비 1,000만", "항암약물치료비"),
    ("19 정액 항암약물치료특약(간편고지형(4), 갱신형) 최초계약 항암방사선약물치료비 1,000만", "항암약물치료비"),
    ("5 정액 항암방사선약물치료비 항암방사선약물치료비 500만", "항암약물방사선"),
    ("9 정액 갱신형 26종 항암방사선및약물치료비(전이포함)(유사암제외)(통합간편가입) 항암방사선약물치료비 1,000만", "항암약물방사선"),
])
def test_anticancer_patterns_no_overcapture_and_split(line, label):
    got = classify_extra(line)
    assert (got[0] if got else None) == label


def test_intensity_proton_and_cart_are_subtracted_from_target_matrix_cell():
    assert TARGET_INCLUDED_LABELS == {"중입자방사선", "세기조절방사선", "양성자방사선", "면역항암치료"}
    matrix = {"표적항암치료": {"by_company": {"3": 5000 * MAN}, "agg": "sum"}}
    extras = {
        "세기조절방사선": {"agg": "sum", "by_company": {"3": 1000 * MAN}, "target_included": {"3": 1000 * MAN}},
        "양성자방사선": {"agg": "sum", "by_company": {"3": 1000 * MAN}, "target_included": {"3": 1000 * MAN}},
        "면역항암치료": {"agg": "sum", "by_company": {"3": 1000 * MAN}, "target_included": {"3": 1000 * MAN}},
    }
    m, x, dedup = _apply_exclusions(matrix, extras)
    assert m["표적항암치료"]["by_company"]["3"] == 2000 * MAN and dedup["subtracted_total"] == 3000 * MAN
    rows = build_v2_rows(m, x)
    assert _row(rows, "cancer_drug_targeted")["summary"] == 2000 * MAN
    assert _row(rows, "radio_imrt")["summary"] == 1000 * MAN and _row(rows, "radio_proton")["summary"] == 1000 * MAN
    assert _row(rows, "cancer_drug_immune")["summary"] == 1000 * MAN


# ── C. 다빈치 3분류 ────────────────────────────────────────────────────────
@pytest.mark.parametrize("name, label", [
    ("다빈치로봇암수술비(연간1회한,특정암제외)", "다빈치(일반암)"),
    ("다빈치로봇암수술비(연간1회한,특정암)", "다빈치(특정암)"),
    ("다빈치로봇수술비(전립선암)", "다빈치(전립선)"),
    ("갑상선암다빈치로봇수술비", "다빈치(갑상선)"),
    ("다빈치로봇수술비", DAVINCI_UNKNOWN_LABEL),
])
def test_davinci_label_never_guesses(name, label):
    assert davinci_label(name) == label
    assert classify_extra(f"1 정액 {name} 암수술 100만")[0] == label


def test_davinci_rows_split_and_matrix_cancer_surgery_is_subtracted():
    matrix = {"암수술": {"by_company": {"1": 600 * MAN}, "agg": "sum"}}
    extras = {
        "다빈치(일반암)": {"agg": "sum", "by_company": {"1": 500 * MAN}, "matrix_included": {"암수술": {"1": 500 * MAN}}},
        "다빈치(특정암)": {"agg": "sum", "by_company": {"1": 100 * MAN}, "matrix_included": {"암수술": {"1": 100 * MAN}}},
        DAVINCI_UNKNOWN_LABEL: {"agg": "sum", "by_company": {"2": 300 * MAN}, "needs_review": True},
    }
    m, x, dedup = _apply_exclusions(matrix, extras)
    assert m["암수술"]["by_company"]["1"] == 0 and dedup["subtracted_total"] == 600 * MAN
    rows = build_v2_rows(m, x)
    assert _row(rows, "cancer_surgery")["summary"] == 0
    dav = _row(rows, "cancer_surgery_davinci")
    assert dav["by_company"] == {"1": 500 * MAN, "2": 300 * MAN} and dav["needs_review"] is True  # 미상은 확인 필요 표기
    assert _row(rows, "cancer_surgery_davinci_specific")["summary"] == 100 * MAN
    assert resolve("다빈치(전립선)").row_id == resolve("다빈치(갑상선)").row_id == "cancer_surgery_davinci_specific"


# ── D. 분배 배선 ──────────────────────────────────────────────────────────
CIRC = """169 갱신형 특정순환계질환 통합치료비
1) 보험기간 중 특정순환계질환의 진단 및 치료를 위한 [특정순환계질환 통합치료(검사)]를 받은 경우
- MRI촬영(급여) : 5만원 (연간 1회한)
2) 주요치료 항목별로 아래의 금액을 지급
- 혈전용해치료 : 1,000만원 (연간 1회한)
- 종합병원 중환자실치료 : 1,000만원 (연간 1회한)
- 수술 : 수술 1회당 1,000만원
5천만원 12,084
※ 위1)~3)까지 합산하여 연간 총 지급액은 5,000만원 한도
176 갱신형 혈전용해치료비Ⅱ(뇌졸중)
1천만원 470"""

CANCER = """117 갱신형 암 통합치료비(실속형)(암중점치료기관(상급종합병원 포함))
암 통합치료보장개시일 이후 통합치료항목별로 연간 1회한 약관에서 정한 금액을 지급
※ 연간 총 지급액은 가입금액을 한도로 함
치료비 3천만원 10,685
124 갱신형 항암중입자방사선치료비
5천만원 1,700
암 통합치료비(실속형)(암중점치료기관(상급종합병원))특약 안내사항
통합치료항목 통합치료항목별 대상질병 지급횟수 지급금액
▶암(유사암제외) 수술 암(유사암제외) 수술 1회당 500만원
▶유사암 수술 유사암 수술 1회당 100만원
주요 ▶암(유사암제외) 항암방사선치료 암(유사암제외) 연간 1회한 500만원
치료 ▶기타피부암 및 갑상선암 항암방사선치료 기타피부암 및 갑상선암 연간 1회한 100만원
▶암(유사암제외) 항암약물치료 암(유사암제외) 연간 1회한 500만원
비급여 비급여(전액본인부담 포함) 표적항암약물허가치료 연간 1회한 500만원
치료 비급여(전액본인부담 포함) 항암양성자방사선치료 연간 1회한 500만원
연간 총 지급액 한도 3천만원"""


def test_circulatory_q8_body_and_surgery_split_limit_is_not_an_item():
    riders = extract_integrated_treatments(CIRC.splitlines())
    r = next(x for x in riders if "순환계" in x["name"])
    assert r["mode"] == MODE_Q8 and r["body_amount"] == 5000 * MAN
    assert [it["name"] for it in r["items"]] == ["MRI촬영(급여)", "혈전용해치료", "종합병원 중환자실치료", "수술"]
    assert any("합산하여" in l for l in r["limit_lines"])           # 한도 문구는 항목이 아니다
    entries = distribution_entries(riders, _entry)
    got = {(e["kb_name"], e["amount"]) for e in entries}
    assert got == {("순환계 치료비", 5000 * MAN), ("뇌혈관수술", 1000 * MAN), ("심혈관수술", 1000 * MAN)}  # 288 Q8 실증값
    assert all(e["origin"] == ORIGIN_DISTRIBUTED and e["merge_rule"] == "sum" for e in entries)


def test_cancer_integrated_table_is_q8_items_routed_body_to_appendix():
    riders = extract_integrated_treatments(CANCER.splitlines())
    r = next(x for x in riders if x["no"] == "117")
    assert r["mode"] == MODE_Q8 and r["body_amount"] == 3000 * MAN
    entries = distribution_entries(riders, _entry)
    got = {e["kb_name"]: e["amount"] for e in entries}
    minor_total = sum(e["amount"] for e in entries if e["kb_name"] == "유사암수술")
    assert got["암수술"] == 500 * MAN and minor_total == 200 * MAN
    assert got["항암방사선치료비"] == 500 * MAN and got["항암약물치료비"] == 500 * MAN
    assert got["표적항암치료"] == 500 * MAN and got["양성자방사선"] == 500 * MAN
    body = next(k for k in got if "연간 총 지급 한도" in k)
    assert got[body] == 3000 * MAN and resolve(body).kind == "appendix"       # 본체(한도)는 비고 보존
    assert route_item("기타피부암 및 갑상선암 항암방사선치료") == "유사암수술"  # Human ⑥ 확정
    assert r["unrouted"] == []


def test_major_treatment_mode_when_amounts_equal_or_missing():
    lines = ["117 갱신형 암 통합치료비(실속형)", "치료비 3천만원 10,685", "999 갱신형 다른담보", "1천만원 1"]
    riders = extract_integrated_treatments(lines)
    assert riders[0]["mode"] == MODE_MAJOR and "미기재" in riders[0]["basis"]
    got = {e["kb_name"]: e["amount"] for e in distribution_entries(riders, _entry)}
    assert got == {"암수술": 3000 * MAN, "항암약물치료비": 3000 * MAN, "항암방사선치료비": 3000 * MAN}
    same = ["169 갱신형 특정순환계질환 통합치료비", "- 혈전용해치료 : 500만원", "- 수술 : 수술 1회당 500만원", "5천만원 1"]
    assert extract_integrated_treatments(same)[0]["mode"] == MODE_MAJOR


def test_q8_without_surgery_item_leaves_r16_r17_blank_with_review_flag():
    lines = ["169 갱신형 특정순환계질환 통합치료비", "- 혈전용해치료 : 1,000만원", "- MRI촬영 : 5만원", "5천만원 1"]
    riders = extract_integrated_treatments(lines)
    entries = distribution_entries(riders, _entry)
    assert {e["kb_name"] for e in entries} == {"순환계 치료비"}
    assert "수술 항목 없음" in riders[0]["needs_review"]


def test_merge_keeps_rep_for_direct_dups_and_sums_distributed_coexistence():
    direct_a = _entry("뇌혈관수술", 300 * MAN, "수술", "sum", "text", "131대질병수술비(뇌혈관질환) 3백만원")
    direct_b = _entry("뇌혈관수술", 300 * MAN, "수술", "sum", "text", "상세면 중복 3백만원")   # 리스트/상세 중복 → rep
    dist_1 = {**_entry("뇌혈관수술", 1000 * MAN, "수술", "sum", "분배:Q8(#169):수술", "수술"), "origin": ORIGIN_DISTRIBUTED, "merge_rule": "sum"}
    dist_2 = {**_entry("암수술", 500 * MAN, "암", "sum", "분배:Q8(#117)", "x"), "origin": ORIGIN_DISTRIBUTED, "merge_rule": "sum"}
    dist_3 = {**_entry("암수술", 1000 * MAN, "암", "sum", "분배:Q8(#119)", "y"), "origin": ORIGIN_DISTRIBUTED, "merge_rule": "sum"}
    merged = {e["kb_name"]: e for e in _merge_entries([direct_a, direct_b, dist_1, dist_2, dist_3])}
    assert merged["뇌혈관수술"]["amount"] == 1300 * MAN and merged["뇌혈관수술"]["origin"] == "direct+distributed"
    assert "text + 분배:Q8" in merged["뇌혈관수술"]["source"]
    assert merged["암수술"]["amount"] == 1500 * MAN and merged["암수술"]["origin"] == ORIGIN_DISTRIBUTED


def test_parse_proposal_text_carries_origin_and_integrated_metadata():
    text = "(무)메리츠 The좋은알파Plus종합보장보험\n월납 77,470원\n" + CIRC + "\n271 갱신형 131대질병수술비(뇌혈관질환) 3백만원 1,209\n"
    r = parse_proposal_text(text, "t.pdf")
    by = {c["kb_name"]: c for c in r["coverages"]}
    assert by["뇌혈관수술"]["amount"] == 1300 * MAN and by["뇌혈관수술"]["origin"] == "direct+distributed"
    assert by["순환계 치료비"]["origin"] == ORIGIN_DISTRIBUTED
    meta = r["metadata"]["integrated_treatments"]
    assert meta and meta[0]["mode"] == MODE_Q8 and any(it["route"] for it in meta[0]["items"])


@pytest.mark.skipif(not (REAL / "20260805_오현지님_보장분석.pdf").exists(), reason="실 PDF 없음(gitignore 폴더)")
def test_real_alpha_plus_proposal_matches_288_q8_evidence():
    import glob
    from coverage.service import analyze_kb_coverage

    prop = sorted(glob.glob(str(REAL / "오현지_77,476원*가입제안서*.pdf")))
    if not prop:
        pytest.skip("제안서 없음")
    parsed = parse_proposal_pdf(Path(prop[0]).read_bytes(), "alpha.pdf")
    by = {c["kb_name"]: c for c in parsed["coverages"]}
    # 288 §3-2 Q8 실증: 본체 5,000만 → 순환계 치료비 / 내역 "수술 1회당 1,000만" → r16·r17 (+ 직접 특약 300만 = sum 1,300만)
    assert by["순환계 치료비"]["amount"] == 5000 * MAN
    assert by["뇌혈관수술"]["amount"] == 1300 * MAN and by["심혈관수술"]["amount"] == 1300 * MAN
    modes = {t["no"]: t["mode"] for t in parsed["metadata"]["integrated_treatments"]}
    assert modes == {"117": MODE_Q8, "119": MODE_Q8, "169": MODE_Q8}   # 안내사항 표 실측 — 금액 상이
    assert by["암수술"]["amount"] == 1500 * MAN and by["다빈치(일반암)"]["amount"] == 1000 * MAN
    assert by["중입자방사선"]["amount"] == 5000 * MAN                    # 신설 매칭 규칙(#124)
    parsed["proposal_id"] = "P1"
    analysis = analyze_kb_coverage((REAL / "20260805_오현지님_보장분석.pdf").read_bytes())
    res = build_after_analysis(analysis, {"existing": [], "proposals": [parsed]})
    rows = res["after"]["before"]["coverages"]
    assert _row(rows, "surgery_cerebral")["by_company"]["P1"] == 1300 * MAN
    assert _row(rows, "circulatory_treatment")["by_company"]["P1"] == 5000 * MAN
    assert _row(rows, "cancer_surgery_davinci")["by_company"]["P1"] == 1000 * MAN


# ── E. 합성 라벨 ─────────────────────────────────────────────────────────
def test_vascular_surgery_label_split_only_when_unambiguous():
    assert classify_extra("1 정액 뇌혈관수술비 질병수술 100만")[0] == "뇌혈관수술비"
    assert classify_extra("1 정액 특정심장혈관수술비 질병수술 100만")[0] == "심혈관수술비"
    assert classify_extra("1 정액 심·뇌혈관수술비 질병수술 100만")[0] == "심·뇌혈관수술비"   # 둘 다 → 비고 보존
    assert resolve("뇌혈관수술비").row_id == "surgery_cerebral" and resolve("심혈관수술비").row_id == "surgery_cardiac"
    assert resolve("심·뇌혈관수술비").kind == "appendix"
    assert anticancer_label("항암방사선약물치료비") == "항암약물방사선" and resolve("항암약물방사선").kind == "appendix"


def test_split_labels_land_on_v2_rows_via_build_before():
    raw = {
        "customer": {"name": "t", "age": 40, "sex": "여자"},
        "contracts": [{"idx": 1, "insurer": "손보", "product": "p", "contract_date": "2024-01-01", "pay_cycle": "월납",
                       "pay_years": 20, "pay_months": 240, "maturity": "100세", "monthly_premium": 10000}],
        "matrix": {}, "diagnosis": {}, "notes": {}, "warnings": [],
        "extra": {"항암약물치료비": {"agg": "sum", "by_company": {"1": 1000 * MAN}},
                  "항암방사선치료비": {"agg": "sum", "by_company": {"1": 1000 * MAN}},
                  "항암약물방사선": {"agg": "sum", "by_company": {"1": 500 * MAN}}},
    }
    before = build_before(raw, today="2026-08-18")
    assert _row(before["coverages"], "cancer_drug")["summary"] == 1000 * MAN
    assert _row(before["coverages"], "cancer_radiation")["summary"] == 1000 * MAN
    assert next(c for c in before["coverages"] if c["kb_name"] == "항암약물방사선")["summary"] == 500 * MAN


# ── F. 2열 라벨 (★BOHUMFIT-296: 292 F 헤더 행 제거 · 행명이 열 순서 표현) ──────────────
def test_excel_no_dual_header_row_and_caregiver_order_unified():
    import io
    import openpyxl
    from tests.test_export_v2_layout_291 import RICH, RICH_EXTRA, _result

    result = _result(RICH, RICH_EXTRA)
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))
    for sheet in ("컨설팅 전", "컨설팅 후"):
        ws = wb[sheet]
        # ★296: 2열 헤더 행 제거 — 컨설팅 전/후도 최종 시트와 동일 좌표(DATA_ROW0 + 순서), 헤더 삽입 0.
        assert track_row_of("tier_surgery_1") == DATA_ROW0 + ROW_INDEX["tier_surgery_1"]
        assert track_row_of("surgery_disease") == DATA_ROW0 + ROW_INDEX["surgery_disease"]
        assert ws.cell(row=track_row_of("tier_surgery_1"), column=3).value == KB_COVERAGES_V2[ROW_INDEX["tier_surgery_1"]].display
        # `2열 병기 (좌 | 우)` 라벨 셀·`질병|상해` 헤더 행이 없어야 한다.
        for r in range(DATA_ROW0, DATA_ROW0 + len(KB_COVERAGES_V2)):
            assert ws.cell(row=r, column=3).value != "2열 병기 (좌 | 우)"
        assert ws.page_setup.fitToWidth == 1
    assert DUAL_ORDER["caregiver"] == ("disease", "injury") == PDF_DUAL_ORDER["caregiver"]
    assert all(DUAL_ORDER[f"tier_surgery_{t}"] == ("disease", "injury") for t in range(1, 6))
    wsf = wb["최종"]
    assert wsf.cell(row=DATA_ROW0 + ROW_INDEX["tier_surgery_1"], column=3).value == KB_COVERAGES_V2[ROW_INDEX["tier_surgery_1"]].display
