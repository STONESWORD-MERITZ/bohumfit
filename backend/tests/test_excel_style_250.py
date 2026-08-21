# -*- coding: utf-8 -*-
"""BOHUMFIT-250 회귀 — 엑셀 시각 디벨롭(값 불변·스타일 존재·면 전용 색 계약). 익명 합성."""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from coverage.aggregator import build_before
from coverage.excel_style import GRAY_SOFT, EMERALD, FILL_ONLY_COLORS, GREENTEA, LIME, WHITE
from coverage.export_excel import build_workbook_bytes
from tests.excel_v2_layout import COL_SUM, SHEET_BEFORE, company_col, row_of  # 291

MAN = 10_000


def _analysis() -> dict:
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [{
            "idx": 1, "insurer": "가나손보", "product": "합성", "contract_date": "2024-01-01",
            "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
            "monthly_premium": 50_000,
        }],
        "matrix": {
            "일반사망" if False else "질병사망": {"by_company": {"1": 5000 * MAN}},
            "암진단금": {"by_company": {"1": 3000 * MAN}},
        },
        "diagnosis": {}, "notes": {},
        "extra": {
            "순환계 치료비": {"agg": "sum", "by_company": {"1": 1000 * MAN}},
            "일반종수술 5종(표준환산)": {"agg": "sum", "by_company": {"1": 500 * MAN}, "estimated": True},
        },
        "warnings": [],
    }
    before = build_before(raw, today="2026-07-27")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def _label_cells(ws) -> dict:
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cells.setdefault(cell.value, cell)
    return cells


def test_style_presence_and_values_unchanged():
    """헤더=에메랄드+흰 글자·대분류 선두 행=그린 티·특수 행(순환계 치료비)=라임 — 값은 집계 그대로.
    291: 시트 `컨설팅 전`(49행 양식). 우측 고객정보 패널(구 O열)은 수기표에 없어 폐기."""
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(_analysis())))
    ws = wb[SHEET_BEFORE]
    header = ws.cell(row=2, column=company_col(0))          # 회사명 헤더
    assert header.fill.fgColor.rgb == EMERALD and header.font.color.rgb == WHITE
    title = ws.cell(row=2, column=2)
    assert title.fill.fgColor.rgb == EMERALD and title.font.color.rgb == WHITE
    # 대분류 선두 행(암 진단비(일반암) = 암 대분류 첫 행): 값 셀에 그린 티 면.
    cancer_row = row_of("cancer_general")
    assert any(ws.cell(row=cancer_row, column=col).fill.fgColor.rgb == GREENTEA
               for col in range(COL_SUM, ws.max_column))
    # ★BOHUMFIT-302b: 순환계 치료비는 **회색 헤더**(`sum_excluded`)로 바뀌었다 — 라임 특수 강조 → 그레이.
    #   대분류 선두 강조(그린 티)는 회색이 아닌 첫 행(뇌 혈관 질환)이 받는다.
    circ_row = row_of("circulatory_treatment")
    assert any(ws.cell(row=circ_row, column=col).fill.fgColor.rgb == GRAY_SOFT
               for col in range(COL_SUM, ws.max_column))
    # 값 검증(스타일 계층이 값을 바꾸지 않음): 질병사망 5,000·암진단 3,000·순환계 1,000(만원).
    assert ws.cell(row=row_of("death_disease"), column=COL_SUM).value == 5000
    assert ws.cell(row=cancer_row, column=COL_SUM).value == 3000
    assert ws.cell(row=circ_row, column=COL_SUM).value == 1000
    # 인쇄 폭 맞춤(15계약 대응 — 사양 결정 3 잠정).
    assert ws.page_setup.fitToWidth == 1


def test_fill_only_colors_never_used_as_font_color():
    """★면 전용 색(라임·그린 티)은 폰트 색으로 사용 금지 — 전 시트 전 셀 검사."""
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(_analysis())))
    for sheet_name in wb.sheetnames:
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                color = getattr(cell.font.color, "rgb", None)
                assert color not in FILL_ONLY_COLORS, (sheet_name, cell.coordinate)
