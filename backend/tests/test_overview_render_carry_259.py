# -*- coding: utf-8 -*-
"""BOHUMFIT-259 회귀 — overview 엑셀 회사 열 렌더 + [후] 회사별 이월.

★계약: ①가드 기준을 "overview 여부" → "by_company 유무"로 전환 ②귀속된 overview 행은
해지가 회사 단위로 반영(249 정본 carry 경로) ③미귀속·부분 귀속은 종전대로 합계만 + 경고
④summary·총액·표준 문서 경로 불변. 익명 합성 픽스처(PII 0).
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from coverage.aggregator import (
    OVERVIEW_CANCEL_WARNING,
    build_before,
    carry_coverage_row,
    overview_rows_need_cancel_warning,
)
from coverage.compare import build_after_analysis
from coverage.export_excel import FORM_ITEMS, build_workbook_bytes

MAN = 10_000


def _ov_row(name, summary, by_company, agg="sum"):
    return {"kb_name": name, "kb_group": "x", "group12": "사망", "agg": agg,
            "summary": summary, "by_company": dict(by_company), "enrolled": True,
            "overview": True}


# ── carry_coverage_row: 귀속 여부에 따른 이월 규칙 ──────────────────────────────
def test_attributed_overview_row_reflects_cancellation():
    """★귀속된 overview 행 — 해지된 계약 열만 빠지고 합계가 재집계된다."""
    row = _ov_row("상해사망", 3000 * MAN, {"1": 1000 * MAN, "2": 2000 * MAN})
    carried = carry_coverage_row(row, kept_ids={"2"}, known_ids={"1", "2"})
    assert carried["by_company"] == {"2": 2000 * MAN}
    assert carried["summary"] == 2000 * MAN
    assert carried["enrolled"] is True
    assert carried["overview"] is True     # 출처 표식 보존


def test_attributed_overview_row_unchanged_when_no_cancel():
    """해지 0이면 회사합=합계 보장에 따라 [전]과 완전히 동일(전=후)."""
    row = _ov_row("상해사망", 3000 * MAN, {"1": 1000 * MAN, "2": 2000 * MAN})
    carried = carry_coverage_row(row, kept_ids={"1", "2"}, known_ids={"1", "2"})
    assert carried["by_company"] == {"1": 1000 * MAN, "2": 2000 * MAN}
    assert carried["summary"] == 3000 * MAN


def test_unattributed_overview_row_keeps_summary_only():
    """미귀속 overview 행은 종전 규칙 — 해지해도 합계 유지(반영 불가)."""
    row = _ov_row("상해사망", 3000 * MAN, {})
    carried = carry_coverage_row(row, kept_ids=set(), known_ids={"1", "2"})
    assert carried["by_company"] == {}
    assert carried["summary"] == 3000 * MAN     # 소실 0(246/249 정책)
    assert carried["enrolled"] is True


def test_unknown_key_still_carried_for_attributed_row():
    """'?' 키는 해지 대상이 아니므로 이월된다(253 모델 유지)."""
    row = _ov_row("상해사망", 3000 * MAN, {"1": 1000 * MAN, "?": 2000 * MAN})
    carried = carry_coverage_row(row, kept_ids=set(), known_ids={"1"})
    assert carried["by_company"] == {"?": 2000 * MAN}
    assert carried["summary"] == 2000 * MAN


# ── 해지 경고: 귀속되면 불필요 ─────────────────────────────────────────────────
def test_cancel_warning_only_for_unattributed_rows():
    assert overview_rows_need_cancel_warning([_ov_row("상해사망", 100, {})]) is True
    assert overview_rows_need_cancel_warning([_ov_row("상해사망", 100, {"1": 100})]) is False
    # 부분 귀속이면 경고 대상(미귀속 행이 남아 있음).
    assert overview_rows_need_cancel_warning(
        [_ov_row("상해사망", 100, {"1": 100}), _ov_row("질병사망", 50, {})]) is True
    assert overview_rows_need_cancel_warning([]) is False


def _analysis(attributed: bool, partial: bool = False):
    """overview 문서 합성 — attributed면 by_company 채움(partial이면 일부만)."""
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [
            {"idx": 1, "insurer": "합성손보1", "product": "합성상품1호", "contract_date": "2021-01-01",
             "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
             "monthly_premium": 10_000},
            {"idx": 2, "insurer": "합성손보2", "product": "합성상품2호", "contract_date": "2022-01-01",
             "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
             "monthly_premium": 20_000},
        ],
        "matrix": {
            "상해사망": {"summary": 3000 * MAN, "overview": True,
                       "by_company": {"1": 1000 * MAN, "2": 2000 * MAN} if attributed else {}},
            "질병사망": {"summary": 1000 * MAN, "overview": True,
                       "by_company": {} if partial else ({"1": 1000 * MAN} if attributed else {})},
        },
        "diagnosis": {}, "notes": {}, "extra": {}, "warnings": [],
    }
    before = build_before(raw, today="2026-07-29")
    return {"before": before,
            "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def _ws(analysis, plan=None):
    result = build_after_analysis(analysis, plan or {"existing": [], "proposals": []})
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))
    return wb["비교분석표"], result


def _row_of(item):
    return 10 + FORM_ITEMS.index(item)


# ── 렌더 가드 전환 ────────────────────────────────────────────────────────────
def test_attributed_overview_renders_company_columns():
    """★259 A: 귀속된 overview는 표준 문서와 동일하게 [전]·[후] 회사 열을 전개한다."""
    ws, result = _ws(_analysis(attributed=True))
    assert ws.cell(row=4, column=2).value == "합성손보1"
    assert ws.cell(row=4, column=3).value == "합성손보2"
    assert ws.cell(row=5, column=2).value == "합성상품1호"     # 2단 헤더(252/254)
    n = 2
    col_bsum = 2 + n
    col_asum = col_bsum + 4
    col_a0 = col_asum + 1
    row = _row_of("상해사망")
    assert [ws.cell(row=row, column=2 + i).value for i in range(n)] == [1000, 2000]
    assert ws.cell(row=row, column=col_bsum).value == 3000
    # [후] 회사 열도 전개(해지 0 → 전=후).
    assert ws.cell(row=4, column=col_a0).value == "합성손보1"
    assert [ws.cell(row=row, column=col_a0 + i).value for i in range(n)] == [1000, 2000]
    assert ws.cell(row=row, column=col_asum).value == 3000


def test_unattributed_overview_keeps_summary_only():
    """★가드 전환 회귀: by_company가 비면 종전대로 회사 열 미생성(합계만)."""
    ws, _res = _ws(_analysis(attributed=False))
    assert ws.cell(row=4, column=2).value == "합 계"      # 회사 열 0 → 2열이 합계
    labels = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "합성손보1" not in labels
    assert ws.cell(row=_row_of("상해사망"), column=2).value == 3000


def test_partially_attributed_overview_keeps_summary_only():
    """부분 귀속(빈 행 혼재)은 빈 회사 열 오독을 피하려 합계만 유지한다."""
    ws, _res = _ws(_analysis(attributed=True, partial=True))
    labels = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "합성손보1" not in labels
    assert ws.cell(row=4, column=2).value == "합 계"


def test_cancellation_drops_only_that_company_column():
    """★259 B: 해지 1건 — [후] 회사 열에서 해당 계약만 빠지고 합계가 재집계된다."""
    analysis = _analysis(attributed=True)
    ws, result = _ws(analysis, {"existing": [{"contract_idx": 1, "disposition": "cancel"}],
                                "proposals": []})
    n, m = 2, 1
    col_bsum = 2 + n
    col_asum = col_bsum + 4
    col_a0 = col_asum + 1
    row = _row_of("상해사망")
    assert ws.cell(row=4, column=col_a0).value == "합성손보2"      # 남은 계약만
    assert ws.cell(row=row, column=col_a0).value == 2000
    assert ws.cell(row=row, column=col_asum).value == 2000         # 합계 재집계
    assert ws.cell(row=row, column=col_bsum).value == 3000         # [전]은 불변
    # 귀속됐으므로 "합계형 문서는 해지 반영 불가" 경고가 붙지 않는다.
    messages = [c.get("message") for c in result["comparison"].get("cautions") or []]
    assert OVERVIEW_CANCEL_WARNING not in messages + (result.get("warnings") or [])


def test_unattributed_cancellation_still_warns():
    """미귀속 overview + 해지 → 보존+경고 정책 유지(246)."""
    analysis = _analysis(attributed=False)
    _ws_, result = _ws(analysis, {"existing": [{"contract_idx": 1, "disposition": "cancel"}],
                                  "proposals": []})
    messages = [c.get("message") for c in result["comparison"].get("cautions") or []]
    assert OVERVIEW_CANCEL_WARNING in messages + (result.get("warnings") or [])


def test_unknown_column_renders_for_attributed_overview():
    """귀속된 overview 행에 '?'가 남으면 "계약 미확인" 열을 노출한다(252 가드 일관 적용)."""
    analysis = _analysis(attributed=True)
    from tests.v2names import v2name
    for row in analysis["before"]["coverages"]:
        if row["kb_name"] == v2name("상해사망"):  # 290: V2 표시명
            row["by_company"] = {"1": 1000 * MAN, "?": 2000 * MAN}
    ws, _res = _ws(analysis)
    labels = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "계약 미확인" in labels
