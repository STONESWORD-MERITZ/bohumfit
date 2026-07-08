from __future__ import annotations

import io
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

from coverage.compare import build_after_analysis
from coverage.export_excel import build_workbook_bytes
from coverage.export_pdf import build_coverage_html

sys.path.insert(0, os.path.dirname(__file__))
from test_coverage_compare_188 import _analysis, _plan  # noqa: E402


def _report() -> dict:
    report = build_after_analysis(_analysis(), _plan())
    report["report_cover"] = {
        "customer_name": "홍길동",
        "ga_name": "리뷰온에셋",
        "planner_name": "김설계",
        "written_date": "2026-07-08",
    }
    return report


def test_pdf_report_uses_six_step_order() -> None:
    html = build_coverage_html(_report(), datetime(2026, 7, 8))

    markers = [
        "① 표지",
        "② 컨설팅 전 계약",
        "③ 신규가입 제안서",
        "④ 최종 전 VS 후 — 특약별 보장 비교",
        "⑤ 최종 전 VS 후 — 회사별 보장 세부",
        "⑥ 컨설팅 전 진단 세부",
    ]
    positions = [html.index(marker) for marker in markers]
    assert positions == sorted(positions)
    assert "해지" in html
    assert "PDF 업로드 슬롯" in html
    assert "BOHUMFIT-193" in html


def test_pdf_splits_rider_and_company_comparison_pages() -> None:
    html = build_coverage_html(_report(), datetime(2026, 7, 8))

    rider_pos = html.index("④ 최종 전 VS 후 — 특약별 보장 비교")
    company_pos = html.index("⑤ 최종 전 VS 후 — 회사별 보장 세부")
    detail_pos = html.index("⑥ 컨설팅 전 진단 세부")
    assert rider_pos < company_pos < detail_pos
    assert "월납입보험료" in html
    assert "20,000원 절감" in html
    assert "특약별 보장 비교 · 개선 담보" in html
    assert "후 합산/대표" in html


def test_excel_uses_same_six_step_sheet_order() -> None:
    workbook = load_workbook(io.BytesIO(build_workbook_bytes(_report())))

    assert workbook.sheetnames == [
        "① 표지",
        "② 전 계약",
        "③ 신규제안",
        "④ 전후 특약별",
        "⑤ 전후 회사별",
        "⑥ 전 진단세부",
    ]
    contract_values = [cell.value for row in workbook["② 전 계약"].iter_rows() for cell in row if cell.value is not None]
    proposal_values = [cell.value for row in workbook["③ 신규제안"].iter_rows() for cell in row if cell.value is not None]
    compare_values = [cell.value for row in workbook["④ 전후 특약별"].iter_rows() for cell in row if cell.value is not None]

    assert "해지" in contract_values
    assert "PDF 업로드 슬롯은 BOHUMFIT-193 파서 연결 전까지 수기 입력 경로와 연결됩니다." in proposal_values
    assert "④ 최종 전 VS 후 - 특약별 보장 비교" in compare_values
    assert -20_000 in compare_values
