from __future__ import annotations

import io
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

from coverage.compare import build_after_analysis
from coverage.export_excel import build_workbook_bytes
from coverage.export_pdf import build_coverage_html

sys.path.insert(0, os.path.dirname(__file__))
from test_coverage_compare_188 import _analysis, _plan  # noqa: E402


def _report() -> dict:
    report = build_after_analysis(_analysis(), _plan())
    report["report_cover"] = {
        "customer_name": "홍길동",
        "insurance_age": "45세",
        "age_change_date": "2026-08-01",
        "ga_name": "리뷰온에셋",
        "planner_name": "김설계",
        "written_date": "2026-07-08",
    }
    return report


def test_pdf_cover_renders_fit_original_cover_fields_and_masks_customer() -> None:
    html = build_coverage_html(_report(), datetime(2026, 7, 8))

    assert "보장분석 리포트" in html
    assert "FIT 보장분석" in html
    assert "GA LOGO" not in html
    assert '<div class="ga-logo-slot">리뷰온에셋</div>' in html
    assert "홍*동" in html
    assert "홍길동" not in html
    for value in ("45세", "2026-08-01", "리뷰온에셋", "김설계", "2026-07-08"):
        assert value in html
    assert "#084734" in html
    assert "보험 모집·중개·상품추천·가입권유" in html


def test_pdf_comparison_has_three_axes_and_group_expansion() -> None:
    html = build_coverage_html(_report(), datetime(2026, 7, 8))

    assert "월납입보험료" in html
    assert "150,000원 → 130,000원" in html
    assert "20,000원 절감" in html
    assert "총납입보험료" in html
    assert "2,400,000원 절감" in html
    assert "대분류별 보장 변화" in html
    assert "특약별 보장금액 비교" in html
    assert "암진단" in html and "수술비" in html


def test_excel_compare_sheet_includes_before_after_premium_and_group_summary() -> None:
    # BOHUMFIT-248 P2: 엑셀 산출을 비분양식 3시트로 정본화 — 검증 의도를 신 양식 등가로 갱신.
    workbook = load_workbook(io.BytesIO(build_workbook_bytes(_report())))
    cover_values = [cell.value for row in workbook["표지(세로)"].iter_rows() for cell in row if cell.value is not None]
    cover_text = " ".join(str(v) for v in cover_values)
    assert "보험 보장 분석 리포트" in cover_text
    assert "리뷰온에셋" in cover_text            # report_cover 소속(GA) — 직급 행에 반영

    b_vals = [cell.value for row in workbook["컨설팅 전"].iter_rows() for cell in row if cell.value is not None]
    a_vals = [cell.value for row in workbook["컨설팅 후"].iter_rows() for cell in row if cell.value is not None]
    assert 150_000 in b_vals and 130_000 in a_vals   # 전/후 월납(원 단위)
    final = workbook["최종"]
    fvals = [cell.value for row in final.iter_rows() for cell in row if cell.value is not None]
    assert -20_000 in fvals                      # 차액 = 후−전
    # 291: 종합 판정 블록 = 케스케이드 17행(구 "암"·"심장말기" 키 없음)
    assert "뇌초기" in fvals and "심장중기" in fvals and "다빈치(일반암)" in fvals
    assert "심장말기" not in fvals
