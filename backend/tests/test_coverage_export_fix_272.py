"""BOHUMFIT-272 — 보장분석 산출물 결함 수정.

★결함 A(최우선): 엑셀 종합비교 '후' 열에 '전' 값이 찍히던 참조 오류.
  실사용 동선의 [후] payload는 클라이언트가 `{...analysis.before}` 스프레드로 만들어
  `stage_totals`가 [전] 값 그대로 남는다(해지가 반영되는 것은 coverages뿐).
  254가 `_yn_flags`에 같은 처방을 했는데 `_stage_map`만 빠져 있었다.
★결함 C: 총납입 정의 2종(계약별 실납입 / 20년 환산)에 라벨을 붙여 서로 모순돼 보이지 않게 한다.
"""
import io
from pathlib import Path

import openpyxl
import pytest

from coverage.aggregator import build_before, build_final, compute_stage_totals
from coverage.compare import build_after_analysis
from coverage.export_excel import MONTHS_20Y, build_workbook_bytes
from tests.excel_v2_layout import SHEET_FINAL, workbook  # 291

ROOT = Path(__file__).resolve().parents[2]
PDF_DIR = ROOT / "보장분석" / "비교분석표"
STANDARD_PDF = next(PDF_DIR.glob("*-INPUT.pdf"), None)


# ── 결함 A: 참조 소스 ──────────────────────────────────────────────────────
def _stage_block(ws) -> dict:
    """291 `최종` 시트 종합 판정 블록(라벨 → (전, 후)) — 헤더 아래 케스케이드 행."""
    out = {}
    start = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.startswith("종합 판정"):
            start = r + 1
            break
    assert start, "종합 판정 블록이 없다"
    r = start
    while r <= ws.max_row:
        v = ws.cell(row=r, column=2).value
        if v in (None, "비고", "특이사항") or (isinstance(v, str) and v.startswith("보장금액 단위")):
            break
        out[v] = (ws.cell(row=r, column=6).value, ws.cell(row=r, column=8).value)
        r += 1
    return out


def test_stage_block_recomputes_from_coverages_not_stale_payload():
    """★payload의 `stage_totals`가 오염돼 있어도 엑셀 종합 블록은 coverages에서 재파생한다."""
    from coverage.aggregator import build_before
    from coverage.compare import build_after_analysis

    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [{"idx": 1, "insurer": "합성", "product": "합성", "contract_date": "2024-01-01",
                       "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
                       "monthly_premium": 10_000}],
        "matrix": {"뇌혈관질환": {"by_company": {"1": 10_000_000}}},
        "diagnosis": {}, "notes": {}, "extra": {}, "warnings": [],
    }
    before = build_before(raw, today="2026-08-17")
    analysis = {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}
    result = build_after_analysis(analysis, {"existing": [], "proposals": []})
    result["after"]["before"]["stage_totals"] = {"뇌초기": 999_999_999}   # ★낡은 값
    block = _stage_block(workbook(result)[SHEET_FINAL])
    expected = compute_stage_totals(result["after"]["before"]["coverages"])
    assert set(block) == set(expected) and len(block) == 17
    for key, value in expected.items():
        assert block[key][1] == round(value / 10_000), key
    assert block["뇌초기"][1] != 999_999_999


@pytest.mark.skipif(
    STANDARD_PDF is None,
    reason="실 PDF 정본 없음(로컬 전용 · PII로 커밋 불가)",
)
def test_excel_stage_matches_pdf_source_when_before_differs_from_after():
    """★전≠후 케이스에서 엑셀 종합 블록(17행) == PDF가 쓰는 소스(재계산 값)."""
    from coverage.aggregator import build_before, build_final
    from coverage.compare import build_after_analysis
    from coverage.parser import parse_document

    raw = parse_document(STANDARD_PDF.read_bytes())
    before = build_before(raw, today="2026-07-29")
    final = build_final(before, raw.get("diagnosis") or {})
    analysis = {"before": before, "final": final, "warnings": []}
    cancels = [{"contract_idx": c["idx"], "disposition": "cancel"} for c in before["contract_list"][:3]]
    result = build_after_analysis(analysis, {"existing": cancels, "proposals": []})
    result["after"]["before"]["stage_totals"] = dict(before["stage_totals"])  # ★클라이언트 스프레드 재현(오염)
    expected_b = compute_stage_totals(before["coverages"])
    expected_a = compute_stage_totals(result["after"]["before"]["coverages"])
    block = _stage_block(workbook(result)[SHEET_FINAL])
    assert list(block) == list(expected_b) and len(block) == 17
    same = 0
    for key in expected_b:
        assert abs((block[key][0] or 0) - round(expected_b[key] / 10_000)) <= 1, key
        assert abs((block[key][1] or 0) - round(expected_a[key] / 10_000)) <= 1, key
        same += block[key][0] == block[key][1]
    # 해지 3건이 어떤 체인 합계는 실제로 낮춘다 — 전부 전=후이면 오염값이 찍힌 것이다.
    assert same < len(block)


# ── 결함 C: 총납입 라벨 ────────────────────────────────────────────────────
def test_pdf_labels_total_payment_definition():
    """PDF 총납입 카드가 **무엇을 뜻하는지** 밝힌다(계약별 납입기간 반영)."""
    src = (ROOT / "backend" / "coverage" / "export_pdf.py").read_text(encoding="utf-8")
    assert src.count("총납입보험료(계약별 납입기간 반영)") == 2
    # 정의 없는 옛 라벨이 남아 있지 않다.
    assert '<div class="k">총납입보험료</div>' not in src
    assert '<div class="k">후 총납입</div>' not in src


def test_excel_keeps_20y_label_and_formula():
    """엑셀 20년 지표는 라벨·산식 그대로다(261 확정분 — 값 변경 금지)."""
    src = (ROOT / "backend" / "coverage" / "export_excel.py").read_text(encoding="utf-8")
    assert "20년 납부 시 총납입 차액" in src
    assert MONTHS_20Y == 240


def test_two_total_payment_concepts_are_distinguishable():
    """두 개념이 **서로 다른 라벨**을 갖는다 — 같은 이름으로 충돌하지 않는다."""
    excel = (ROOT / "backend" / "coverage" / "export_excel.py").read_text(encoding="utf-8")
    pdf = (ROOT / "backend" / "coverage" / "export_pdf.py").read_text(encoding="utf-8")
    assert "20년 납부 시 총납입 차액" in excel
    assert "총납입보험료(계약별 납입기간 반영)" in pdf
    # 20년 환산 라벨이 PDF에 잘못 들어가 있지 않다(개념 혼선 방지).
    assert "20년 납부 시" not in pdf


# ── 보호 영역 ──────────────────────────────────────────────────────────────
def test_272_does_not_touch_protected_modules():
    """★`pipeline/`·`filters.py`에 272 흔적이 없다."""
    assert "272" not in (ROOT / "backend" / "filters.py").read_text(encoding="utf-8")
    for path in (ROOT / "backend" / "pipeline").glob("*.py"):
        assert "BOHUMFIT-272" not in path.read_text(encoding="utf-8"), path.name


def test_form_schema_is_v2_49_rows():
    """291: 엑셀 양식 = V2 49행 스키마(대분류 11) — 구 FORM_ITEMS(35)+YN(5)는 폐기됐다."""
    from coverage.constants import GROUP12_V2, KB_COVERAGES_V2
    import coverage.export_excel as ex

    assert len(KB_COVERAGES_V2) == 52 and len(GROUP12_V2) == 11  # 296: +3행
    assert not hasattr(ex, "FORM_ITEMS") and not hasattr(ex, "YN_ROWS") and not hasattr(ex, "STAGE_ROWS")
