# -*- coding: utf-8 -*-
from __future__ import annotations

import io

from openpyxl import load_workbook

from coverage.compare import build_after_analysis, compare_before_after
from coverage.export_excel import build_workbook_bytes
from coverage.export_pdf import build_coverage_html


def _analysis() -> dict:
    before = {
        "customer": {"name": "테스트", "age": 40, "sex": "남"},
        "premium": {"monthly_total": 150_000, "paid_total": 18_000_000, "currency": "KRW"},
        "companies": [
            {
                "idx": 1,
                "insurer": "A보험",
                "product": "기존 주계약",
                "contract_date": "2020-01-01",
                "pay_cycle": "월납",
                "pay_years": 10,
                "pay_months": 120,
                "maturity": "90세",
                "monthly_premium": 100_000,
                "paid_total": 12_000_000,
                "remark": None,
            },
            {
                "idx": 2,
                "insurer": "B보험",
                "product": "해지 후보",
                "contract_date": "2021-01-01",
                "pay_cycle": "월납",
                "pay_years": 10,
                "pay_months": 120,
                "maturity": "90세",
                "monthly_premium": 50_000,
                "paid_total": 6_000_000,
                "remark": None,
            },
        ],
        "coverages": [
            {
                "kb_name": "일반사망",
                "kb_group": "사망",
                "group12": "사망",
                "agg": "sum",
                "summary": 100_000_000,
                "by_company": {"1": 100_000_000, "2": None},
                "enrolled": True,
            },
            {
                "kb_name": "암진단",
                "kb_group": "암",
                "group12": "암",
                "agg": "sum",
                "summary": None,
                "by_company": {"1": None, "2": None},
                "enrolled": False,
            },
            {
                "kb_name": "수술비",
                "kb_group": "수술",
                "group12": "수술",
                "agg": "sum",
                "summary": 10_000_000,
                "by_company": {"1": 10_000_000, "2": None},
                "enrolled": True,
            },
        ],
    }
    before["contract_list"] = before["companies"]
    final = {
        "premium": before["premium"],
        "coverages": [
            {
                "group12": "사망",
                "kb_name": "일반사망",
                "agg": "sum",
                "value": 100_000_000,
                "recommended": 100_000_000,
                "gap": 0,
                "status": "충분",
            },
            {
                "group12": "암",
                "kb_name": "암진단",
                "agg": "sum",
                "value": None,
                "recommended": 50_000_000,
                "gap": -50_000_000,
                "status": "미가입",
            },
            {
                "group12": "수술",
                "kb_name": "수술비",
                "agg": "sum",
                "value": 10_000_000,
                "recommended": 20_000_000,
                "gap": -10_000_000,
                "status": "부족",
            },
        ],
        "rollup_by_group12": [],
    }
    return {"before": before, "final": final, "warnings": []}


def _plan() -> dict:
    return {
        "version": 1,
        "source": "coverage-remodel",
        "existing": [
            {"contract_idx": 1, "disposition": "keep"},
            {"contract_idx": 2, "disposition": "cancel"},
        ],
        "proposals": [
            {
                "proposal_id": "P1",
                "insurer": "신규보험",
                "product": "암수술 보완",
                "monthly_premium": 30_000,
                "pay_cycle": "월납",
                "pay_months": 120,
                "maturity": "100세",
                "coverages": [
                    {"kb_name": "암진단", "amount": 50_000_000},
                    {"kb_name": "수술비", "amount": 10_000_000},
                ],
            }
        ],
    }


def test_after_builder_compares_monthly_paid_status_and_improvements() -> None:
    result = build_after_analysis(_analysis(), _plan())

    assert result["before"]["premium"]["monthly_total"] == 150_000
    assert result["after"]["before"]["premium"]["monthly_total"] == 130_000
    assert result["comparison"]["premium"]["delta_monthly"] == -20_000
    assert result["comparison"]["premium"]["delta_paid_total"] == -2_400_000

    after_by_name = {row["kb_name"]: row for row in result["after"]["final"]["coverages"]}
    assert after_by_name["암진단"]["status"] == "충분"
    assert after_by_name["수술비"]["status"] == "충분"
    assert after_by_name["암진단"]["value"] == 50_000_000
    assert after_by_name["수술비"]["value"] == 20_000_000

    comparison = result["comparison"]
    assert comparison["summary"]["improved_count"] == 2
    assert comparison["summary"]["missing_to_sufficient"] == 1
    assert comparison["summary"]["short_to_sufficient"] == 1
    changes = {row["kb_name"]: row["status_change"] for row in comparison["coverages"]}
    assert changes["암진단"] == "미가입 -> 충분"
    assert changes["수술비"] == "부족 -> 충분"


def test_compare_before_after_can_be_called_directly() -> None:
    result = build_after_analysis(_analysis(), _plan())
    comparison = compare_before_after(_analysis()["final"], result["after"]["final"])

    assert comparison["premium"]["before_monthly"] == 150_000
    assert comparison["premium"]["after_monthly"] == 130_000
    assert comparison["summary"]["improved_count"] == 2


def test_excel_adds_after_compare_and_summary_sheets() -> None:
    result = build_after_analysis(_analysis(), _plan())
    workbook = load_workbook(io.BytesIO(build_workbook_bytes(result)))

    assert workbook.sheetnames == ["① 표지", "② 전 계약", "③ 신규제안", "④ 전후 특약별", "⑤ 전후 회사별"]
    compare = workbook["④ 전후 특약별"]
    values = [cell.value for row in compare.iter_rows() for cell in row if cell.value is not None]
    assert "전 보장금액" in values
    assert "후 보장금액" in values
    assert "미가입 -> 충분" not in values
    assert "부족 -> 충분" not in values
    assert -20_000 in values


def test_pdf_html_adds_customer_compare_page_with_brand_and_disclaimer() -> None:
    html = build_coverage_html(build_after_analysis(_analysis(), _plan()))

    assert "컨설팅 전 VS 후 요약" in html
    assert "특약별 보장금액 비교" in html
    assert "부족·미가입 → 충분" not in html
    assert "보험 모집·중개·상품추천·가입권유" in html
    assert "#084734" in html
    for old in ("#15663D", "#2E6B3E", "#145C2A"):
        assert old not in html
