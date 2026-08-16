# -*- coding: utf-8 -*-
"""BOHUMFIT-261 회귀 — 엑셀 표지 리디자인 · 20년 총납입 차액 · PDF 고령 가독성.

★불변 계약: 담보 값·회사합=합계·40행·[후] 이월은 그대로(표시·레이아웃 계층만).
익명 합성 픽스처(PII 0).
"""
from __future__ import annotations

import io
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from coverage.aggregator import build_before
from coverage.compare import build_after_analysis
from coverage.excel_style import EMERALD, WHITE
from coverage.export_excel import MONTHS_20Y, build_workbook_bytes
from tests.excel_v2_layout import COL_SUM, SHEET_BEFORE, SHEET_FINAL, company_col, row_of  # 291
from coverage.export_pdf import COMPANY_CHUNK, build_coverage_html

MAN = 10_000
GEN = datetime(2026, 7, 31)




def _analysis(companies: int = 2, monthly=(10_000, 20_000)) -> dict:
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [
            {"idx": i, "insurer": f"합성손보{i}", "product": f"합성상품{i}호",
             "contract_date": "2024-01-01", "pay_cycle": "월납", "pay_years": 20,
             "pay_months": 240, "maturity": "100세", "monthly_premium": monthly[i - 1]}
            for i in range(1, companies + 1)
        ],
        "matrix": {"질병사망": {"by_company": {"1": 3000 * MAN, "2": 2000 * MAN}}},
        "diagnosis": {}, "notes": {}, "extra": {}, "warnings": [],
    }
    before = build_before(raw, today="2026-07-31")
    return {"before": before,
            "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []},
            "report_cover": {"planner_name": "김설계", "ga_name": "합성지사",
                             "planner_tel": "010-0000-0000"}}


def _wb(analysis, plan=None):
    result = build_after_analysis(analysis, plan or {"existing": [], "proposals": []})
    return openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result, generated_at=GEN))), result


# ── P1 표지 ────────────────────────────────────────────────────────────────
def test_cover_has_brand_band_title_date_and_planner_block():
    wb, _res = _wb(_analysis())
    ws = wb["표지(세로)"]
    band = ws.cell(row=2, column=2)
    assert "BohumFit" in str(band.value) and "보험핏" in str(band.value)
    assert band.fill.fgColor.rgb == EMERALD and band.font.color.rgb == WHITE  # 브랜드 밴드
    assert ws.cell(row=7, column=2).value == "홍길동 님을 위한"
    assert ws.cell(row=8, column=2).value == "보험 보장 분석 리포트"
    assert ws.cell(row=11, column=2).value == "작성일  2026-07-31"
    labels = [ws.cell(row=r, column=3).value for r in range(13, 17)]
    assert labels == ["소속(GA)", "설계사명", "연락처", "E-MAIL"]
    assert ws.cell(row=13, column=7).value == "합성지사"      # 제공값은 채움
    assert ws.cell(row=14, column=7).value == "김설계"
    # 미제공 항목은 값 없이 테두리만 있는 기입란(openpyxl은 빈 문자열을 None으로 저장).
    assert ws.cell(row=16, column=7).value is None
    assert ws.cell(row=16, column=7).border.bottom.style is not None
    assert "약관과 증권" in str(ws.cell(row=20, column=2).value)  # 하단 고지
    # A4 세로 1장 인쇄 정합.
    assert ws.page_setup.orientation == "portrait"
    assert ws.page_setup.fitToWidth == 1 and ws.page_setup.fitToHeight == 1


def test_cover_falls_back_when_no_planner_data():
    analysis = _analysis()
    analysis["report_cover"] = {}
    wb, _res = _wb(analysis)
    ws = wb["표지(세로)"]
    assert all(ws.cell(row=r, column=7).value is None for r in range(13, 17))
    assert ws.cell(row=7, column=2).value == "홍길동 님을 위한"


# ── P2 20년 총납입 차액 ────────────────────────────────────────────────────
def test_twenty_year_delta_row_added_and_matches_formula():
    """월납 차액 × 240 — 291 `최종` 시트: 4행 월납(전/후/차액) · 6행 20년 총납입(전/후/차액) 병기."""
    analysis = _analysis()
    plan = {"existing": [{"contract_idx": 1, "disposition": "cancel"}], "proposals": []}
    wb, result = _wb(analysis, plan)
    ws = wb[SHEET_FINAL]
    before_monthly = result["before"]["premium"]["monthly_total"]
    after_monthly = result["after"]["before"]["premium"]["monthly_total"]
    delta = after_monthly - before_monthly
    assert delta == -10_000                                   # 계약1(월 10,000) 해지
    assert ws.cell(row=4, column=6).value == before_monthly
    assert ws.cell(row=4, column=8).value == after_monthly
    assert ws.cell(row=4, column=10).value == delta            # 월납 차액
    assert ws.cell(row=6, column=6).value == before_monthly * MONTHS_20Y
    assert ws.cell(row=6, column=10).value == delta * MONTHS_20Y == -2_400_000  # 20년 총납입 차액


def test_twenty_year_delta_color_follows_direction():
    """절감=에메랄드 / 증가=앰버(250 규칙) — 291: 종합 판정·담보 기대효과 셀 색(후−전 부호)."""
    from coverage.excel_style import AMBER_TX

    analysis = _analysis()
    wb, _res = _wb(analysis, {"existing": [{"contract_idx": 1, "disposition": "cancel"}], "proposals": []})
    ws = wb[SHEET_FINAL]
    assert ws.cell(row=4, column=10).font.color.rgb == EMERALD
    assert ws.cell(row=6, column=10).font.color.rgb == EMERALD
    # 계약1 해지 → 질병사망 3,000만 감소(앰버) — 기대효과 열
    assert ws.cell(row=row_of("death_disease"), column=10).value == -3000
    assert ws.cell(row=row_of("death_disease"), column=10).font.color.rgb == AMBER_TX
    # 증가 케이스: 신규 제안으로 담보가 늘면 에메랄드.
    analysis2 = _analysis()
    plan = {"existing": [], "proposals": [
        {"proposal_id": "P1", "insurer": "신규", "product": "신규안", "monthly_premium": 50_000,
         "pay_months": 240, "maturity": "100세",
         "coverages": [{"kb_name": "질병사망", "amount": 1000 * MAN, "group12": "사망", "agg": "sum"}]}]}
    wb2, _res2 = _wb(analysis2, plan)
    assert wb2[SHEET_FINAL].cell(row=4, column=10).font.color.rgb == AMBER_TX
    assert wb2[SHEET_FINAL].cell(row=6, column=10).font.color.rgb == AMBER_TX
    cell = wb2[SHEET_FINAL].cell(row=row_of("death_disease"), column=10)
    assert cell.value == 1000 and cell.font.color.rgb == EMERALD


# ── P3 PDF 가독성 ──────────────────────────────────────────────────────────
def test_pdf_font_scaled_up_for_senior_readers():
    _wb_, result = _wb(_analysis())
    html = build_coverage_html(result, GEN)
    assert "font-size: 13.5pt" in html          # 본문(구 10pt)
    assert "padding: 7px 8px; font-size: 12pt" in html  # 표 셀(구 9pt)
    assert "h2 {{ font-size: 16pt" in html or "h2 { font-size: 16pt" in html
    assert "font-size: 9pt; }" not in html      # 구 표 셀 크기 잔존 0


def test_pdf_sections_start_on_new_pages_with_repeating_headers():
    _wb_, result = _wb(_analysis())
    html = build_coverage_html(result, GEN)
    assert "page-break-before: always" in html
    assert "first-section" in html               # 첫 섹션은 예외(빈 페이지 방지)
    assert "display: table-header-group" in html  # 표 헤더 매 페이지 반복
    assert "page-break-inside: avoid" in html     # 행 쪼개짐 방지


def test_pdf_company_columns_are_chunked():
    """★회사 열 분할 — 폰트 확대 후 다열 표가 A4를 넘치지 않도록 5개씩 끊는다."""
    _wb_, result = _wb(_analysis(companies=2))
    html = build_coverage_html(result, GEN)
    assert html.count('<th rowspan="2">담보</th>') == 1        # 2계약 → 1개 표
    assert 'class="chunk-caption"' not in html               # 단일 표는 캡션 마크업 없음

    many = _analysis(companies=2)
    companies = many["before"]["contract_list"]
    for idx in range(3, 13):                                   # 12계약으로 확장
        extra = dict(companies[0])
        extra.update({"idx": idx, "insurer": f"합성손보{idx}", "product": f"합성상품{idx}호"})
        companies.append(extra)
    many["before"]["companies"] = companies
    _wb2_, result2 = _wb(many)
    html2 = build_coverage_html(result2, GEN)
    expected = -(-12 // COMPANY_CHUNK)
    assert html2.count('<th rowspan="2">담보</th>') == expected
    assert html2.count('class="chunk-caption"') == expected   # 묶음별 캡션


def test_amount_cells_unchanged_by_presentation_changes():
    """★값 불변: 담보 금액·회사합=합계는 표시 개편 후에도 그대로."""
    wb, result = _wb(_analysis())
    ws = wb[SHEET_BEFORE]
    rows = {c["row_id"]: c for c in result["before"]["coverages"] if c.get("row_id")}
    row = row_of("death_disease")
    assert ws.cell(row=row, column=company_col(0)).value == 3000
    assert ws.cell(row=row, column=company_col(1)).value == 2000
    assert ws.cell(row=row, column=COL_SUM).value == 5000            # 합계 열
    assert rows["death_disease"]["summary"] == 5000 * MAN
