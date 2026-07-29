# -*- coding: utf-8 -*-
"""BOHUMFIT-252 회귀 — 시트2 회사별 열 전개(전/후) + 2단 헤더 + [후] 신규 골격.

익명 합성 픽스처. ★값·집계 불변 계약: 렌더 계층만 검증 — by_company/summary 값이
시트 셀에 그대로(만원) 옮겨지고, 회사별 열 합 = 합계 열(대사 0)임을 고정한다.
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from coverage.aggregator import build_before
from coverage.compare import build_after_analysis
from coverage.excel_style import EMERALD, WHITE
from coverage.export_excel import FORM_ITEMS, build_workbook_bytes

MAN = 10_000


def _analysis(companies: int = 2) -> dict:
    """계약 2개·담보별 by_company 채움(만원 정수 배수 — 반올림 편차 0 설계)."""
    contracts = [
        {
            "idx": i, "insurer": f"합성손보{i}", "product": f"합성상품{i}호",
            "contract_date": f"202{i}-01-01", "pay_cycle": "월납", "pay_years": 20,
            "pay_months": 240, "maturity": "100세", "monthly_premium": 10_000 * i,
        }
        for i in range(1, companies + 1)
    ]
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": contracts,
        "matrix": {
            "질병사망": {"by_company": {"1": 3000 * MAN, "2": 2000 * MAN}},
            "암진단금": {"by_company": {"1": 1000 * MAN}},
            "질병후유장해": {"by_company": {"2": 500 * MAN}},
        },
        "diagnosis": {}, "notes": {},
        "extra": {},
        "warnings": [],
    }
    before = build_before(raw, today="2026-07-27")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def _overview_analysis() -> dict:
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [{
            "idx": i, "insurer": f"합성손보{i}", "product": f"합성상품{i}호",
            "contract_date": "2024-01-01", "pay_cycle": "월납", "pay_years": 20,
            "pay_months": 240, "maturity": "100세", "monthly_premium": 10_000,
        } for i in range(1, 3)],
        "matrix": {
            "상해사망": {"summary": 30000 * MAN, "by_company": {}, "overview": True},
            "암진단금": {"summary": 10000 * MAN, "by_company": {}, "overview": True},
        },
        "diagnosis": {}, "notes": {}, "extra": {}, "warnings": [],
    }
    before = build_before(raw, today="2026-07-27")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def _ws(result: dict):
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))
    return wb["비교분석표"]


def _item_row(item: str) -> int:
    return 10 + FORM_ITEMS.index(item)


def test_two_tier_header_company_and_product():
    """252 A: 4행=회사명·5행=상품명(에메랄드+흰 글자) — 보험료 중복 기재는 50행 한 곳만."""
    result = build_after_analysis(_analysis(), {"existing": [], "proposals": []})
    ws = _ws(result)
    assert ws.cell(row=4, column=2).value == "합성손보1"
    assert ws.cell(row=4, column=3).value == "합성손보2"
    assert ws.cell(row=5, column=2).value == "합성상품1호"
    assert ws.cell(row=5, column=3).value == "합성상품2호"
    assert ws.cell(row=5, column=2).fill.fgColor.rgb == EMERALD
    assert ws.cell(row=5, column=2).font.color.rgb == WHITE
    # 회사별 월납은 한 곳에만 — 5행에 숫자 없음(254 개정2: 위치는 상단 9행 "월보험료").
    assert not isinstance(ws.cell(row=5, column=2).value, (int, float))
    assert ws.cell(row=9, column=2).value == 10_000
    assert ws.cell(row=9, column=3).value == 20_000


def test_company_columns_match_payload_and_sum_to_total():
    """★회사별 열 값 = payload by_company(만원) · 회사별 합 = 합계 열(전·후 대사 0)."""
    result = build_after_analysis(_analysis(), {"existing": [], "proposals": []})
    ws = _ws(result)
    before_rows = {c["kb_name"]: c for c in result["before"]["coverages"]}
    after_rows = {c["kb_name"]: c for c in result["after"]["before"]["coverages"]}
    n = 2
    col_bsum = 2 + n            # [전] 합계
    col_asum = col_bsum + 4     # 담보명 3열 뒤 [후] 합계
    col_a0 = col_asum + 1
    for item in ("질병사망", "암진단금", "질병후유장해"):
        row = _item_row(item)
        b = before_rows[item]
        a = after_rows[item]
        for idx in range(n):
            key = str(idx + 1)
            expect_b = (b.get("by_company") or {}).get(key)
            expect_a = (a.get("by_company") or {}).get(key)
            assert ws.cell(row=row, column=2 + idx).value == (None if expect_b is None else round(expect_b / MAN))
            assert ws.cell(row=row, column=col_a0 + idx).value == (None if expect_a is None else round(expect_a / MAN))
        # 회사별 합 = 합계 열(만원 정수 배수 픽스처 — 반올림 편차 0).
        b_cells = [ws.cell(row=row, column=2 + idx).value or 0 for idx in range(n)]
        a_cells = [ws.cell(row=row, column=col_a0 + idx).value or 0 for idx in range(n)]
        assert sum(b_cells) == (ws.cell(row=row, column=col_bsum).value or 0)
        assert sum(a_cells) == (ws.cell(row=row, column=col_asum).value or 0)


def test_after_new_contract_skeleton_column():
    """252 B: 제안 미착수 — [후] 말미에 '신규 설계 반영 대상' 골격 열(값 전부 공란)."""
    result = build_after_analysis(_analysis(), {"existing": [], "proposals": []})
    ws = _ws(result)
    col_new = (2 + 2) + 4 + 1 + 2  # [전]합계+담보명3+[후]합계+[후]2열 다음
    assert ws.cell(row=4, column=col_new).value == "신규 설계 반영 대상"
    for offset in range(len(FORM_ITEMS)):
        assert ws.cell(row=10 + offset, column=col_new).value is None
    assert ws.cell(row=50, column=col_new).value is None


def test_no_skeleton_when_proposal_contract_present():
    """제안 계약(신규제안)이 [후]에 이미 있으면 골격 열을 만들지 않는다."""
    result = build_after_analysis(_analysis(), {"existing": [], "proposals": []})
    result["after"]["before"]["contract_list"] = list(result["after"]["before"]["contract_list"]) + [{
        "idx": "P1", "insurer": "신규제안", "product": "신규설계", "contract_date": None,
        "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
        "monthly_premium": 30_000, "remark": "신규제안", "consulting_status": "신규제안",
    }]
    ws = _ws(result)
    labels = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "신규 설계 반영 대상" not in labels
    assert "신규제안" in labels  # 제안 계약이 일반 회사 열로 전개


def test_overview_no_company_columns_and_no_skeleton():
    """246 계약 유지: overview 문서 — 회사 열·골격 열 미생성, 합계 열만."""
    result = build_after_analysis(_overview_analysis(), {"existing": [], "proposals": []})
    ws = _ws(result)
    # [전] 합계=2열·[후] 합계=6열(회사 열 0) — 249 고정 좌표.
    assert ws.cell(row=3, column=2).value == "비교분석 전 보장"
    assert ws.cell(row=4, column=2).value == "합 계"
    assert ws.cell(row=4, column=6).value == "합 계"
    labels = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "신규 설계 반영 대상" not in labels
    row = _item_row("상해사망")
    assert ws.cell(row=row, column=2).value == 30000
    assert ws.cell(row=row, column=6).value == 30000


# ── 252 재개: '?'(계약 미상) 버킷 조건부 "계약 미확인" 열 ─────────────────────────
def _analysis_with_unknown_bucket() -> dict:
    """FORM 담보(항암약물방사선)에 '?' 버킷이 남는 합성 — 동일 보험료 모호 문서 재현."""
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [{
            "idx": 1, "insurer": "합성손보1", "product": "합성상품1호",
            "contract_date": "2024-01-01", "pay_cycle": "월납", "pay_years": 20,
            "pay_months": 240, "maturity": "100세", "monthly_premium": 10_000,
        }],
        "matrix": {"질병사망": {"by_company": {"1": 3000 * MAN}}},
        "diagnosis": {}, "notes": {},
        "extra": {"항암약물방사선": {"agg": "sum", "by_company": {"?": 200 * MAN}}},
        "warnings": [],
    }
    before = build_before(raw, today="2026-07-27")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def test_unknown_bucket_renders_explicit_column_and_reconciles():
    """'?' 잔존 시 "계약 미확인" 열 명시 출력 — 회사열+미확인 = 합계([전]·[후] 이월 포함)."""
    result = build_after_analysis(_analysis_with_unknown_bucket(), {"existing": [], "proposals": []})
    ws = _ws(result)
    n = 1
    col_b_unk = 2 + n                 # [전] 미확인 열(회사 열 뒤·합계 앞)
    col_bsum = col_b_unk + 1
    col_asum = col_bsum + 4
    col_a0 = col_asum + 1
    col_a_unk = col_a0 + n            # [후] 미확인 열
    assert ws.cell(row=4, column=col_b_unk).value == "계약 미확인"
    assert ws.cell(row=4, column=col_a_unk).value == "계약 미확인"
    row = _item_row("항암약물방사선")
    assert ws.cell(row=row, column=col_b_unk).value == 200          # '?' 200만원 → 만원
    assert ws.cell(row=row, column=col_bsum).value == 200           # 합계와 대사 성립
    assert ws.cell(row=row, column=2).value is None                 # 실계약 열은 공란
    assert ws.cell(row=row, column=col_a_unk).value == 200          # [후] 이월('?' 보존 — 246)
    # 신규 골격 열은 미확인 열 뒤로 밀림.
    assert ws.cell(row=4, column=col_a_unk + 1).value == "신규 설계 반영 대상"
    # 보험료 행은 미확인 열 공란(보험료 개념 없음).
    assert ws.cell(row=50, column=col_b_unk).value is None


def test_no_unknown_column_when_bucket_absent():
    """'?' 없으면 미확인 열 미출력(현행 유지) — 실 5케이스 잔존 0(253) 상태의 기본형."""
    result = build_after_analysis(_analysis(), {"existing": [], "proposals": []})
    ws = _ws(result)
    labels = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert "계약 미확인" not in labels


def test_sheet3_design_headers():
    """252 C: 시트3 헤더(리모델링 전/주요보장/리모델링 후/종합/전/비교/후)=에메랄드+흰 글자."""
    result = build_after_analysis(_analysis(), {"existing": [], "proposals": []})
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))
    ws = wb["최종비교분석표"]
    for col, text in ((2, "리모델링 전"), (3, "주요보장"), (6, "리모델링 후"),
                      (8, "종합"), (9, "전"), (10, "비교"), (11, "후")):
        cell = ws.cell(row=4, column=col)
        assert cell.value == text
        assert cell.fill.fgColor.rgb == EMERALD and cell.font.color.rgb == WHITE
