# -*- coding: utf-8 -*-
"""BOHUMFIT-181 보장분석 내보내기(엑셀/PDF) 테스트.

★PII: 픽스처 익명(홍길동)·합성. 엑셀은 openpyxl 왕복 검증, PDF는 HTML 문자열 검증
   (실제 PDF 렌더=Codex 로컬 playwright). 파일명 규칙은 sanitize 로직 검증.
"""
import io
import re

from openpyxl import load_workbook

from coverage.export_excel import build_workbook_bytes
from coverage.export_pdf import build_coverage_html, _fmt_krw

MAN = 10_000
EOK = 100_000_000

ANALYSIS = {
    "before": {
        "customer": {"name": "홍길동", "age": 33, "sex": "남"},
        "premium": {"monthly_total": 573227, "paid_total": 181984128, "currency": "KRW"},
        "companies": [
            {"idx": 6, "insurer": "삼성생명", "product": "퍼펙트통합", "pay_years": 15,
             "maturity": "종신", "monthly_premium": 333600, "paid_total": 60048000, "remark": "계피동일"},
            {"idx": 2, "insurer": "KB손보", "product": "KB Yes365", "pay_years": 25,
             "maturity": "90세", "monthly_premium": 101463, "paid_total": 30438900, "remark": "계피동일"},
        ],
        "coverages": [
            {"kb_name": "상해사망", "kb_group": "사망", "group12": "사망", "agg": "sum",
             "summary": 550000000, "by_company": {"6": 200000000, "2": 200000000}, "enrolled": True},
            {"kb_name": "N대수술비", "kb_group": "기타", "group12": "기타", "agg": "sum",
             "summary": 40900000, "by_company": {"2": 30400000, "6": 10500000}, "enrolled": True},
        ],
    },
    "final": {
        "premium": {"monthly_total": 573227, "paid_total": 181984128},
        "coverages": [
            {"group12": "사망", "kb_name": "상해사망", "agg": "sum", "value": 550000000,
             "recommended": 200000000, "gap": 350000000, "status": "충분"},
            {"group12": "실비", "kb_name": "유사암", "agg": "sum", "value": 10000000,
             "recommended": 20000000, "gap": -10000000, "status": "부족"},
            {"group12": "기타", "kb_name": "N대수술비", "agg": "sum", "value": 40900000,
             "recommended": None, "gap": None, "status": None},
        ],
    },
    "warnings": [],
}


def test_excel_sheets_and_values():
    data = build_workbook_bytes(ANALYSIS)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["최종 보장진단", "전 회사별세부"]
    fin = wb["최종 보장진단"]
    assert fin["B2"].value == 573227 and fin["E2"].value == 181984128
    # BOHUMFIT-237 A: 보장금액 셀은 한글 단위 문자열(만/억은 엑셀 표시 포맷으로 불가 — 문자열 전환).
    vals = [c.value for row in fin.iter_rows(min_row=5) for c in row]
    assert "5억 5,000만원" in vals and "2억원" in vals   # 상해사망 가입/권장
    bef = wb["전 회사별세부"]
    bvals = [c.value for row in bef.iter_rows() for c in row]
    assert "5억 5,000만원" in bvals and "4,090만원" in bvals   # 상해사망 합산 + 기타 N대수술비
    text = " ".join(str(c.value) for row in bef.iter_rows() for c in row if c.value is not None)
    assert "계피동일" in text and "N대수술비" in text


def test_pdf_html_brand_and_values():
    html = build_coverage_html(ANALYSIS)
    assert "#084734" in html                      # 에메랄드
    for old in ("#15663D", "#2E6B3E", "#145C2A"):  # 구 브랜드색
        assert old not in html
    assert "BohumFit" in html and "보험핏" in html and "ㅍ" in html
    assert "삼성생명" in html and "KB Yes365" in html
    assert "333,600원" in html and "101,463원" in html
    assert "컨설팅 전 진단 세부" not in html
    assert "충분" not in html and "부족" not in html


def test_fmt_krw():
    # BOHUMFIT-237 A: 보장금액 한글 단위 표기 — "만원/억원" 접미 통일.
    assert _fmt_krw(550000000) == "5억 5,000만원"
    assert _fmt_krw(200000000) == "2억원"
    assert _fmt_krw(300000) == "30만원"
    assert _fmt_krw(120000000) == "1억 2,000만원"
    assert _fmt_krw(5000) == "5,000원"
    assert _fmt_krw(None) == "-"
    assert _fmt_krw(0) == "0원"


def _safe_name(label, ext):
    """main._coverage_export_names sanitize 로직 재현 검증."""
    safe = re.sub(r"[^가-힣A-Za-z0-9]", "", str(label or ""))[:20] or "고객"
    return f"BohumFit_보장분석_{safe}_20260707.{ext}"


def test_filename_sanitize():
    assert _safe_name("홍길동", "xlsx") == "BohumFit_보장분석_홍길동_20260707.xlsx"
    assert _safe_name("", "pdf") == "BohumFit_보장분석_고객_20260707.pdf"          # 기본값
    assert "/" not in _safe_name("김/철*수?", "xlsx") and "*" not in _safe_name("김/철*수?", "xlsx")
    assert _safe_name("김/철*수?", "xlsx") == "BohumFit_보장분석_김철수_20260707.xlsx"  # 금지문자 제거
