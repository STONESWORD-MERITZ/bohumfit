# -*- coding: utf-8 -*-
"""BOHUMFIT-248 P2/P5 회귀 — 비분양식 엑셀의 QA 발견 결함 고정(익명 합성 픽스처).

P4 QA에서 발견한 결함: overview(계약 0열) 문서에서 부록 라벨 열과 값 열이 동일 열로
겹쳐 담보명이 숫자로 덮임(정보 소실). 수정 후 라벨/값 열 분리를 고정한다.
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from coverage.aggregator import build_before
from coverage.export_excel import build_workbook_bytes
from tests.excel_v2_layout import COL_SUM, SHEET_BEFORE, row_of  # 291

MAN = 10_000


def _analysis(raw_overrides: dict) -> dict:
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [],
        "matrix": {},
        "diagnosis": {},
        "notes": {},
        "extra": {},
        "warnings": [],
    }
    raw.update(raw_overrides)
    before = build_before(raw, today="2026-07-26")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def test_overview_appendix_labels_not_overwritten():
    """overview(계약 열 0)에서도 부록 라벨과 값이 서로 다른 열에 남는다(P4 결함 회귀)."""
    analysis = _analysis({
        "matrix": {
            "상해사망": {"summary": 30000 * MAN, "by_company": {}, "overview": True},
        },
        "extra": {"80%이상 후유장해": {"agg": "sum", "by_company": {"1": 100 * MAN}}},
    })
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(analysis)))
    ws = wb[SHEET_BEFORE]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    # BOHUMFIT-290(S2·Q2): 80% 담보는 이제 후유장해 대분류의 정식 행이라 부록이 아니라 본 표에 실린다.
    #   구 라벨 "80%이상 후유장해"는 V2 표시명 "상해 질병 후 유 장 해 80%"로 보인다(값 100 보존).
    assert "상해 질병 후 유 장 해 80%" in text
    values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert 100 in values                       # 값(만원)도 별도 열에 존재


def test_estimated_note_and_overview_summary_column():
    """238 표준환산 문구 병기 + overview 합계 열 값 기입."""
    analysis = _analysis({
        "matrix": {"암진단금": {"summary": 5000 * MAN, "by_company": {}, "overview": True}},
        "extra": {"일반종수술 5종(표준환산)": {"agg": "sum", "by_company": {"1": 1000 * MAN}, "estimated": True}},
    })
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(analysis)))
    ws = wb[SHEET_BEFORE]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "표준환산" in text  # 291: 하단 안내 "표준환산 종수술은 종별 미확인으로 병합 표기"
    # overview: 계약 열 0 → 합계 열(F). 암진단비(일반암) 값 5,000(만원). 표준환산 5종은 병합 셀 1,000.
    assert ws.cell(row=row_of("cancer_general"), column=COL_SUM).value == 5000
    assert ws.cell(row=row_of("tier_surgery_5"), column=COL_SUM).value == 1000
