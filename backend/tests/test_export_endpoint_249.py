# -*- coding: utf-8 -*-
"""BOHUMFIT-249 회귀 — ★사용자 동선 기준(엔드포인트 POST 경유) export 검증.

배경: 248 P2 검증이 exporter 직접 호출이어서 UI 동선·배포 반영 확인이 누락됐고,
프로덕션 실물에서 구 5시트 + overview [후] 0원이 발견됐다(249 S0 판정: 배포 스테일 +
compare 경로의 246 보정 누락). 이 테스트는 UI가 실제로 때리는 경로를 그대로 고정한다.
익명 합성 픽스처(홍길동)만 사용.
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import pytest

from coverage.aggregator import build_before
from coverage.compare import build_after_analysis

MAN = 10_000


# BOHUMFIT-290(S2): 집계 행이 V2 49행 — 구 이름 조회는 export와 같은 투영(legacy_form_view)으로.
from tests.v2names import legacy_form_view as _view  # noqa: E402
from coverage.v2_mapping import GROUP_APPENDIX_V2 as _APPENDIX  # noqa: E402


@pytest.fixture()
def client():
    from fastapi.testclient import TestClient

    import main

    main.app.dependency_overrides[main.verify_jwt] = lambda: "user-test"
    yield TestClient(main.app)
    main.app.dependency_overrides.clear()


def _overview_analysis() -> dict:
    """E형(239 합계-only) 익명 재현 — [후] 이월 결함의 재발 방지 표적."""
    raw = {
        "customer": {"name": "홍길동", "age": 56, "sex": "여자"},
        "contracts": [{
            "idx": 1, "insurer": "가나손보", "product": "합성", "contract_date": "2024-01-01",
            "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
            "monthly_premium": 4_675_189,
        }],
        "matrix": {
            "상해사망": {"summary": 30000 * MAN, "by_company": {}, "overview": True},
            "암진단금": {"summary": 10000 * MAN, "by_company": {}, "overview": True},
        },
        "diagnosis": {}, "notes": {}, "extra": {}, "warnings": [],
    }
    before = build_before(raw, today="2026-07-27")
    return {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}


def test_excel_endpoint_produces_form_sheets_and_preserves_overview_after(client):
    """UI 동선(POST /coverage/export/excel): 비분양식 3시트 + overview [후] 합계 보존."""
    analysis = _overview_analysis()
    # 서버 권위 [후] 경로(compare.build_after_analysis — 249에서 단일 소스로 보정)로 after 구성.
    result = build_after_analysis(analysis, {"existing": [], "proposals": []})
    response = client.post("/coverage/export/excel", json=result)
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    # ① 291 4시트 — 구 3/5시트 구조는 어떤 payload로도 생성 불가.
    assert wb.sheetnames == ["표지(세로)", "컨설팅 전", "컨설팅 후", "최종"]
    from tests.excel_v2_layout import COL_SUM, ROW_PREMIUM, row_of
    ws_b, ws_a = wb["컨설팅 전"], wb["컨설팅 후"]
    # ② overview: [전]·[후] 합계 열 모두 값 보존(만원). 계약 열 0 → 합계(F열).
    assert ws_b.cell(row=row_of("death_injury"), column=COL_SUM).value == 30000   # [전] 상해사망 3억
    assert ws_a.cell(row=row_of("death_injury"), column=COL_SUM).value == 30000   # ★[후] 보존 — 249 재발 방지
    assert ws_b.cell(row=row_of("cancer_general"), column=COL_SUM).value == 10000  # 암진단 1억
    assert ws_a.cell(row=row_of("cancer_general"), column=COL_SUM).value == 10000
    # ③ 월납 이월(원 단위) — 291: 6행 월납(값 불변).
    assert ws_b.cell(row=ROW_PREMIUM, column=COL_SUM).value == 4_675_189
    assert ws_a.cell(row=ROW_PREMIUM, column=COL_SUM).value == 4_675_189


def test_excel_endpoint_standard_before_after_equal_when_no_cancel(client):
    """표준(계약별 셀) 문서 — 해지 0이면 엔드포인트 산출 [전]=[후]."""
    raw = {
        "customer": {"name": "홍길동", "age": 50, "sex": "남자"},
        "contracts": [{
            "idx": 1, "insurer": "가나손보", "product": "합성", "contract_date": "2024-01-01",
            "pay_cycle": "월납", "pay_years": 20, "pay_months": 240, "maturity": "100세",
            "monthly_premium": 50_000,
        }],
        "matrix": {"질병사망": {"by_company": {"1": 5000 * MAN}}},
        "diagnosis": {}, "notes": {}, "extra": {}, "warnings": [],
    }
    before = build_before(raw, today="2026-07-27")
    analysis = {"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}}
    result = build_after_analysis(analysis, {"existing": [], "proposals": []})
    response = client.post("/coverage/export/excel", json=result)
    assert response.status_code == 200
    from tests.excel_v2_layout import COL_SUM, row_of
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    # 질병사망 행 — 컨설팅 전·후 합계(F열) 동일(해지 0).
    assert wb["컨설팅 전"].cell(row=row_of("death_disease"), column=COL_SUM).value == 5000
    assert wb["컨설팅 후"].cell(row=row_of("death_disease"), column=COL_SUM).value == 5000


def test_server_after_path_carries_overview_and_unknown_keys():
    """서버 권위 [후](compare.build_after_analysis) — 249 단일 소스 보정의 직접 회귀."""
    analysis = _overview_analysis()
    # 계약 미상 키('?') 행 추가.
    analysis["before"]["coverages"].append({
        "kb_name": "재해사망(계약 미확인)", "kb_group": "기타", "group12": "기타", "agg": "sum",
        "summary": 10000 * MAN, "by_company": {"?": 10000 * MAN}, "enrolled": True,
    })
    result = build_after_analysis(analysis, {"existing": [{"contract_idx": 1, "disposition": "cancel"}]})
    rows = _view(result["after"]["before"]["coverages"])
    # overview 행: 전 계약 해지에도 합계 보존 + 경고.
    assert rows["상해사망"]["summary"] == 30000 * MAN
    assert any("합계" in w for w in result.get("warnings", []))
    # '?' 키: 해지와 무관하게 이월.
    assert rows["재해사망(계약 미확인)"]["summary"] == 10000 * MAN
