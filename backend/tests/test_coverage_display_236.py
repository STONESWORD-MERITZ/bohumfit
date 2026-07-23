# -*- coding: utf-8 -*-
"""BOHUMFIT-236 표시 개선 회귀 — 익명 합성 픽스처(홍길동)만 사용.

A 납입완료 판별·월납 병기 / B 계약 번호 숫자 정렬·"계약 N" 헤더 통일 /
E constants 2대주요치료비 패턴. (C/D/E UI는 프런트 테스트가 담당)
"""
from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from coverage.aggregator import _is_paid_up, build_before  # noqa: E402
from coverage.constants import classify_extra  # noqa: E402
from coverage.export_excel import build_workbook_bytes  # noqa: E402
from coverage.export_pdf import build_coverage_html  # noqa: E402

TODAY = "2026-07-21"


def _contract(idx, premium, date="2026-01-01", years=20, cycle="월납", **extra):
    return {
        "idx": idx,
        "insurer": f"보험사{idx}",
        "product": f"상품{idx}",
        "contract_date": date,
        "pay_cycle": cycle,
        "pay_years": years,
        "pay_months": years * 12,
        "maturity": "100세",
        "monthly_premium": premium,
        **extra,
    }


def _raw(contracts):
    return {
        "customer": {"name": "홍길동", "age": 58, "sex": "남자"},
        "contracts": contracts,
        "matrix": {},
        "diagnosis": {},
        "notes": {},
        "extra": {},
        "warnings": [],
    }


# ── A: 납입완료 판별 + 월납 병기 산식 ────────────────────────────────────────
def test_paid_up_detection():
    assert _is_paid_up(_contract(1, 10000, date="2003-03-28", years=15), TODAY) is True
    assert _is_paid_up(_contract(2, 10000, date="2021-02-23", years=20), TODAY) is False
    # 일시납은 즉시 납입완료
    assert _is_paid_up(_contract(3, 10000, date="2025-08-05", years=1, cycle="일시납"), TODAY) is True
    # 경계: 종료일이 오늘과 같으면 완료
    assert _is_paid_up(_contract(4, 10000, date="2016-07-21", years=10), TODAY) is True
    assert _is_paid_up(_contract(5, 10000, date="2016-07-22", years=10), TODAY) is False


def test_monthly_total_dual_values():
    """주값=전체 합산(납입완료 포함), 부값=납입완료 제외 — KB 원본 헤더 산식(234 실측)."""
    before = build_before(
        _raw([
            _contract(1, 100_000, date="2003-01-01", years=15),   # 납입완료
            _contract(2, 200_000, date="2025-01-01", years=20),   # 납입 중
            _contract(3, 999_000, date="2026-02-01", years=1, cycle="일시납"),  # 월납 합산 제외
        ]),
        today=TODAY,
    )
    assert before["premium"]["monthly_total"] == 300_000
    assert before["premium"]["monthly_total_active"] == 200_000
    flags = {c["idx"]: c["paid_up"] for c in before["companies"]}
    assert flags == {1: True, 2: False, 3: True}


# ── B: 계약 번호 숫자 정렬(사전식 1,10,11,…,2 버그 수정) ─────────────────────
def test_companies_sorted_numerically():
    contracts = [_contract(i, 1000 * i) for i in (11, 2, 1, 10)]
    before = build_before(_raw(contracts), today=TODAY)
    assert [c["idx"] for c in before["companies"]] == [1, 2, 10, 11]


# ── B: 엑셀/PDF 헤더 "계약 N" 통일 + A 병기·납입완료 표기 ────────────────────
def test_excel_matrix_header_and_paid_up_label():
    import openpyxl

    contracts = [
        _contract(1, 100_000, date="2003-01-01", years=15),
        _contract(2, 200_000, date="2025-01-01", years=20),
    ]
    before = build_before(_raw(contracts), today=TODAY)
    data = build_workbook_bytes({"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    sheet = wb["전 회사별세부"]
    headers = [cell.value for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str)]
    # BOHUMFIT-240 P1: 매트릭스 헤더가 회사명(고유 보험사 → 그대로). "계약 N"은 fallback 전용.
    assert "보험사1" in headers and "보험사2" in headers
    assert "계약 1" not in headers and "계약 2" not in headers


def test_excel_matrix_header_duplicate_insurer_disambiguated():
    """BOHUMFIT-240 P1: 동일 회사 복수 계약은 '회사명 (1)/(2)'로 구분."""
    import openpyxl

    contracts = [
        {**_contract(1, 100_000), "insurer": "삼성화재"},
        {**_contract(2, 200_000), "insurer": "삼성화재"},
    ]
    before = build_before(_raw(contracts), today=TODAY)
    data = build_workbook_bytes({"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    headers = [cell.value for row in wb["전 회사별세부"].iter_rows() for cell in row if isinstance(cell.value, str)]
    assert "삼성화재 (1)" in headers and "삼성화재 (2)" in headers


def test_pdf_html_dual_premium_and_paid_up_chip():
    contracts = [
        _contract(1, 100_000, date="2003-01-01", years=15),
        _contract(2, 200_000, date="2025-01-01", years=20),
    ]
    before = build_before(_raw(contracts), today=TODAY)
    html = build_coverage_html({"before": before, "final": {"premium": before["premium"], "coverages": [], "rollup_by_group12": []}})
    assert "월납 합계 300,000원" in html
    assert "납입완료 제외 시 200,000원" in html
    assert html.count("납입완료</span>") >= 1  # 완료 계약 배지


# ── E: 2대주요치료비 패턴 자동 분류(뇌·심 계열) ─────────────────────────────
def test_two_major_treatment_pattern():
    got = classify_extra("2대주요치료비(뇌·심)수술비 특정질병수술 500만")
    assert got is not None and got[0] == "2대주요치료비(뇌·심)"
    # N대수술비로 과포섭되지 않는다(236 E — 234 ② 계열 재발 방지)
    got2 = classify_extra("2대주요치료비(혈전용해치료) 기타수술 300만")
    assert got2 is not None and got2[0] == "2대주요치료비(뇌·심)"


# ── F: 실사용 A 케이스 c4 비표준 담보 2종 가시화(간편보험 양식) ──────────────
def test_simplified_plan_riders_classified():
    got = classify_extra("4 정액 간편심사[355(6대)] 상해중환자실입원일당(1-180) 상해중환자실입원일당 20만")
    assert got is not None and got[0] == "중환자실 입원일당"
    # 전각 ％·반각 % 모두 수용
    got2 = classify_extra("1 정액 간편심사[355(6대)] 일반상해 80％이상 후유장해 상해80%이상후유장해 1,000만")
    assert got2 is not None and got2[0] == "80%이상 후유장해"
    got3 = classify_extra("일반상해 80%이상 후유장해 상해후유장해 500만")
    assert got3 is not None and got3[0] == "80%이상 후유장해"
