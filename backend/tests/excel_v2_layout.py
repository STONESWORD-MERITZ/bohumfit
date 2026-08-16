# -*- coding: utf-8 -*-
"""BOHUMFIT-291(S3) — 테스트 공용: 49행 양식 엑셀 좌표 헬퍼.

★엑셀 양식이 비분양식(시트 `비교분석표`·10행 시작·회사 1열)에서 수기표 49행 양식
  (시트 `컨설팅 전`/`컨설팅 후`/`최종`·7행 시작·회사 2열)으로 바뀌었다. 좌표에 묶인 테스트는
  이 헬퍼로 조회하고, **단언 의미(값 불변·회사합=합계·색·가드)는 그대로** 지킨다.
"""
from __future__ import annotations

import io

import openpyxl

from coverage.constants import KB_COVERAGES_V2, SHEET_NAME_AFTER_V2, SHEET_NAME_BEFORE_V2, SHEET_NAME_FINAL_V2
from coverage.export_excel import DATA_ROW0, build_workbook_bytes
from coverage.v2_mapping import ROW_INDEX

SHEET_BEFORE = SHEET_NAME_BEFORE_V2
SHEET_AFTER = SHEET_NAME_AFTER_V2
SHEET_FINAL = SHEET_NAME_FINAL_V2

COL_GROUP = 2      # B 대분류
COL_NAME = 3       # C~E 담보명(병합)
COL_SUM = 6        # F~G 합계(2열)
COL_CO0 = 8        # H~ 회사별 2열
ROW_COMPANY = 2    # 회사명
ROW_PERIOD = 3     # 납기만기
ROW_PRODUCT = 4    # 상품명(4~5 병합)
ROW_PREMIUM = 6    # 월납
LAST_DATA_ROW = DATA_ROW0 + len(KB_COVERAGES_V2) - 1
APPENDIX_ROW0 = LAST_DATA_ROW + 1

# 최종 시트
FINAL_COL_BEFORE = 6
FINAL_COL_AFTER = 8
FINAL_COL_DELTA = 10


def workbook(result: dict):
    return openpyxl.load_workbook(io.BytesIO(build_workbook_bytes(result)))


def row_of(row_id: str) -> int:
    return DATA_ROW0 + ROW_INDEX[row_id]


def company_col(i: int) -> int:
    """i번째(0-base) 회사의 좌측 열."""
    return COL_CO0 + 2 * i


def header_labels(ws) -> list:
    return [ws.cell(row=ROW_COMPANY, column=c).value for c in range(1, ws.max_column + 1)]


def pair_value(ws, row: int, col: int):
    """병합 2열 값(좌 셀) — 2열 병기 행이면 (좌, 우) 튜플."""
    left, right = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col + 1).value
    return (left, right) if right is not None else left


def company_cells(ws, row: int, n: int) -> list:
    """회사별 2열 값을 회사당 하나(좌+우 합)로 — 회사합=합계 대사용."""
    out = []
    for i in range(n):
        col = company_col(i)
        left, right = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col + 1).value
        left = left if isinstance(left, (int, float)) else 0
        right = right if isinstance(right, (int, float)) else 0
        out.append(left + right)
    return out


def sum_cell(ws, row: int):
    left, right = ws.cell(row=row, column=COL_SUM).value, ws.cell(row=row, column=COL_SUM + 1).value
    left = left if isinstance(left, (int, float)) else 0
    right = right if isinstance(right, (int, float)) else 0
    return left + right


def appendix_labels(ws) -> list:
    out = []
    r = APPENDIX_ROW0
    while r <= ws.max_row:
        v = ws.cell(row=r, column=COL_NAME).value
        if ws.cell(row=r, column=COL_GROUP).value not in ("비고", None):
            break
        if v and v != "-":
            out.append(v)
        r += 1
        if ws.cell(row=r, column=COL_GROUP).value is None and ws.cell(row=r, column=COL_NAME).value is None:
            break
    return out
